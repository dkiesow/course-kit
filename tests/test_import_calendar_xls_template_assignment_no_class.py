import sqlite3
from pathlib import Path
from openpyxl import Workbook, load_workbook
import pytest

pytest.importorskip('openpyxl')
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

from scripts.import_calendar_xls import import_calendar_xls


def test_assignment_based_no_class(tmp_path):
    db_path = tmp_path / 'test_no_class.db'
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT,
            short TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            order_index INTEGER,
            unit TEXT,
            reading_list TEXT,
            monday_details TEXT,
            wednesday_details TEXT
        )
    ''')
    conn.commit()
    conn.close()

    # Calendar workbook with assignment tokens 'Holiday' and 'No Class'
    cal_path = tmp_path / 'calendar_nc.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'SP2026'
    ws.append(['Unit', 'Reading', 'Mon', 'Wed', 'Assignments'])
    ws.append(['Unit 1 - 1/19/26', 'Ch 1', '', '', 'Holiday'])
    ws.append(['Unit 2 - 1/26/26', 'Ch 2', '', '', 'No Class'])
    wb.create_sheet('SP2026 Dates')
    wb.save(str(cal_path))

    # Template with a single placeholder row (week/date in col A)
    template_path = tmp_path / 'template_nc.xlsx'
    twb = Workbook()
    tws = twb.active
    tws.title = 'SP2026'
    tws.append(['{DATES}', '{readings}', None, None, '{ASSIGNMENT}'])
    twb.create_sheet('SP2026 Dates')
    twb.save(str(template_path))

    out_path = tmp_path / 'populated_nc.xlsx'

    import_calendar_xls(str(cal_path), db_path=str(db_path), semester='SP2026', dry_run=True, populate_template=str(template_path), populate_output=str(out_path))

    assert out_path.exists()
    wb2 = load_workbook(str(out_path), data_only=True)
    main = wb2['SP2026']

    # Find rows with dates 2026-01-19 and 2026-01-26
    def iso(v):
        import datetime
        if v is None:
            return None
        if hasattr(v, 'date'):
            return v.date().isoformat()
        if isinstance(v, datetime.date):
            return v.isoformat()
        return None

    r1 = r2 = None
    for r in range(1, main.max_row + 1):
        v = main.cell(row=r, column=1).value
        if iso(v) == '2026-01-19':
            r1 = r
        if iso(v) == '2026-01-26':
            r2 = r
    assert r1 and r2

    # Assignments column should have labels and rows should be shaded
    assert main.cell(row=r1, column=5).value == 'Holiday'
    assert main.cell(row=r2, column=5).value == 'No Class'
    assert getattr(main.cell(row=r1, column=1).fill, 'fill_type', None) == 'solid' or getattr(main.cell(row=r1, column=1).fill, 'patternType', None) == 'solid'
    assert getattr(main.cell(row=r2, column=1).fill, 'fill_type', None) == 'solid' or getattr(main.cell(row=r2, column=1).fill, 'patternType', None) == 'solid'
