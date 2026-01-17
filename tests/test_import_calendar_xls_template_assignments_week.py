import sqlite3
import pytest
pytest.importorskip('openpyxl')
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))
from pathlib import Path
from openpyxl import load_workbook, Workbook

from scripts.import_calendar_xls import import_calendar_xls


def test_populate_template_writes_week_and_assignments(tmp_path):
    db_path = tmp_path / "test_template.db"
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT,
            short TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            order_index INTEGER,
            unit TEXT,
            reading_list TEXT,
            monday_details TEXT,
            wednesday_details TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE deck_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            deck_id INTEGER NOT NULL,
            assignment_id INTEGER NOT NULL,
            UNIQUE(deck_id, assignment_id)
        )
    ''')
    # Insert a couple of assignments for SP2026
    c.execute('INSERT INTO assignments (id, semester, name, due_date, short) VALUES (?, ?, ?, ?, ?)',
              (1, 'SP2026', 'Homework One', '2026-01-19', 'HW1'))
    c.execute('INSERT INTO assignments (id, semester, name, due_date, short) VALUES (?, ?, ?, ?, ?)',
              (2, 'SP2026', 'Homework Two', '2026-01-26', 'HW2'))
    conn.commit()
    conn.close()

    # Create a simple calendar workbook with two data rows
    cal_path = tmp_path / 'calendar.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'SP2026'
    # Header row (will be skipped)
    ws.append(['Unit', 'Reading', 'Mon', 'Wed', 'Assignment'])
    # Two rows with Unit and assignment short in column E
    ws.append(['Unit 1 - 1/19/26', 'Ch 1', 'Intro', 'Quiz', 'HW1'])
    ws.append(['Unit 2 - 1/26/26', 'Ch 2', 'Lab', 'Test', 'HW2'])
    wb.create_sheet('SP2026 Dates')
    wb.save(str(cal_path))

    # Create a template workbook with placeholder rows for WEEK/DATE/READINGS/ASSIGNMENT
    template_path = tmp_path / 'template.xlsx'
    twb = Workbook()
    tws = twb.active
    tws.title = 'SP2026'
    # Placeholder rows that the importer should replace
    tws.append(['{WEEK #}', '{readings}', None, None, '{ASSIGNMENT(Short)}'])
    tws.append(['{DATES}', '{readings}', None, None, '{ASSIGNMENT(Short)}'])
    twb.create_sheet('SP2026 Dates')
    twb.save(str(template_path))

    out_path = tmp_path / 'populated.xlsx'

    # Run importer which should populate the template and write out_path
    import_calendar_xls(str(cal_path), db_path=str(db_path), semester='SP2026', dry_run=True, populate_template=str(template_path), populate_output=str(out_path))

    assert out_path.exists()
    wb2 = load_workbook(str(out_path), data_only=True)
    assert 'SP2026' in wb2.sheetnames
    main = wb2['SP2026']
    # Week labels should be written as words into the first placeholder row (col A) and into the first row of the appended block
    assert main.cell(row=1, column=1).value == 'One'
    assert main.cell(row=3, column=1).value == 'Two'
    # Assignments (shorts) should appear in column E for the second and third rows of each populated block
    assert main.cell(row=2, column=5).value == 'HW1'
    assert main.cell(row=4, column=5).value == 'HW2'
