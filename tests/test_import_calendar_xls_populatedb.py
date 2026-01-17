import sqlite3
import pytest
pytest.importorskip('openpyxl')
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))
from openpyxl import load_workbook, Workbook
from scripts.import_calendar_xls import import_calendar_xls


def test_populate_template_from_db_writes_topics(tmp_path):
    db_path = tmp_path / "test_template_db_only.db"
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            unit TEXT,
            reading_list TEXT,
            monday_details TEXT,
            wednesday_details TEXT,
            topic1 TEXT,
            topic2 TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT,
            short TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE deck_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            deck_id INTEGER NOT NULL,
            assignment_id INTEGER NOT NULL,
            UNIQUE(deck_id, assignment_id)
        )
    ''')
    # Insert one deck with topics
    c.execute('INSERT INTO decks (id, presentation_id, week, date, unit, topic1, topic2) VALUES (?, ?, ?, ?, ?, ?, ?)',
              (1, 42, 'Week 1', '2026-01-13', 'Unit 1', 'DB Topic A', 'DB Topic B'))
    conn.commit()
    conn.close()

    # Create template with TOPIC placeholders in main sheet
    template_path = tmp_path / 'template_db_only.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'SP2026'
    ws.append(['{DATES}', '{READINGS}', '{MON}', '{WED}', '{ASSIGNMENT}', '{TOPIC1}', '{TOPIC2}'])
    wb.create_sheet('SP2026 Dates')
    wb.save(str(template_path))

    out_path = tmp_path / 'populated_db_only.xlsx'

    # Run importer with populate-from-db True, restricting to presentation_id 42
    import_calendar_xls(str(template_path), db_path=str(db_path), semester='SP2026', dry_run=True, populate_template=str(template_path), populate_output=str(out_path), populate_from_db=True, presentation_id=42)

    assert out_path.exists()
    wb2 = load_workbook(str(out_path))
    ws2 = wb2['SP2026']
    # TOPIC1 -> column 6, TOPIC2 -> column 7 in the first populated block
    assert ws2.cell(row=1, column=6).value == 'DB Topic A'
    assert ws2.cell(row=1, column=7).value == 'DB Topic B'
