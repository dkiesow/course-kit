import sqlite3
import sys, pathlib
from pathlib import Path
from openpyxl import Workbook, load_workbook
import pytest

pytest.importorskip('openpyxl')
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

from scripts.import_calendar_xls import import_calendar_xls


def test_populate_template_with_no_class_days(tmp_path):
    # Setup minimal DB (no assignments needed for this test)
    db_path = tmp_path / 'test_holidays.db'
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT,
            short TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            order_index INTEGER,
            unit TEXT,
            reading_list TEXT,
            monday_details TEXT,
            wednesday_details TEXT
        )
    ''')
    conn.commit()
    conn.close()

    # Create a minimal template workbook with SP2026 main sheet and SP2026 Dates
    template_path = tmp_path / 'template_holidays.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'SP2026'
    # add a merged header to ensure formatting is preserved
    ws['B1'] = 'Header'
    ws.merge_cells('B1:C1')
    # Create two placeholder rows for dates
    ws['A10'] = '{DATES}'
    ws['B10'] = '{readings}'
    ws['A14'] = '{DATES}'
    ws['B14'] = '{readings}'
    # Create dates sheet
    wb.create_sheet('SP2026 Dates')
    wb.save(str(template_path))

    # Create an input calendar workbook (xls) with two rows: MLK Day and a normal day
    cal_path = tmp_path / 'calendar.xlsx'
    cw = Workbook()
    ws2 = cw.active
    ws2.append(['Unit', 'Reading', 'MON', 'WED', 'Assignments'])
    ws2.append(['Unit 1 - 1/21/26', 'Ch. 1', '', '', ''])  # MLK Day (expected no-class)
    ws2.append(['Unit 2 - 1/26/26', 'Ch. 2', '', '', ''])  # normal
    cw.save(str(cal_path))

    out_path = tmp_path / 'populated_holidays.xlsx'

    # Call importer with a no_class_dates mapping for MLK Day
    no_class_map = {'2026-01-21': 'MLK DAY'}
    import_calendar_xls(str(cal_path), db_path=str(db_path), semester='SP2026', dry_run=True, populate_template=str(template_path), populate_output=str(out_path), no_class_dates=no_class_map)

    # Verify output file exists and has expected markings
    assert out_path.exists()
    wb2 = load_workbook(str(out_path), data_only=True)
    assert 'SP2026' in wb2.sheetnames
    main = wb2['SP2026']

    # Find the row that got populated with 2026-01-21
    import datetime

    def _iso_date_str(v):
        if v is None:
            return None
        if hasattr(v, 'date'):
            return v.date().isoformat()
        if isinstance(v, datetime.date):
            return v.isoformat()
        if isinstance(v, str):
            # try normalize simple tokens like '1/21/26' -> '2026-01-21'
            from scripts.import_calendar_xls import parse_any_date
            return parse_any_date(v) or v
        return None

    found_row = None
    for r in range(1, main.max_row + 1):
        v = main.cell(row=r, column=1).value
        if _iso_date_str(v) == '2026-01-21':
            found_row = r
            break
    assert found_row is not None, 'MLK date not written to main sheet'

    # MLK Day is a Tuesday/Wednesday depending on calendar; check both MON/WED cells for the label
    mon_val = main.cell(row=found_row, column=3).value
    wed_val = main.cell(row=found_row, column=4).value
    assert (mon_val == 'MLK DAY') or (wed_val == 'MLK DAY')

    # Check formatting merge preserved (B1:C1)
    merges = [str(m) for m in main.merged_cells.ranges]
    assert 'B1:C1' in merges


def test_assignment_column_triggers_no_class(tmp_path):
    # Setup DB with a Holiday assignment (short = MLK)
    db_path = tmp_path / 'test_holidays2.db'
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT,
            short TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            order_index INTEGER,
            unit TEXT,
            reading_list TEXT,
            monday_details TEXT,
            wednesday_details TEXT
        )
    ''')
    # Insert Holiday assignment
    c.execute('INSERT INTO assignments (id, semester, name, due_date, short) VALUES (?, ?, ?, ?, ?)',
              (1, 'SP2026', 'Holiday', '2026-01-19', 'MLK'))
    conn.commit()
    conn.close()

    # Template workbook
    template_path = tmp_path / 'template_holiday_assign.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'SP2026'
    # placeholder row
    ws['A10'] = '{DATES}'
    ws['B10'] = '{readings}'
    wb.create_sheet('SP2026 Dates')
    wb.save(str(template_path))

    # Calendar workbook with assignment token 'Holiday'
    cal_path = tmp_path / 'calendar2.xlsx'
    cw = Workbook()
    ws2 = cw.active
    ws2.append(['Unit', 'Reading', 'MON', 'WED', 'Assignments'])
    ws2.append(['Unit 1 - 1/19/26', 'Ch. 1', '', '', 'Holiday'])
    cw.save(str(cal_path))

    out_path = tmp_path / 'populated_holiday_assign.xlsx'

    import_calendar_xls(str(cal_path), db_path=str(db_path), semester='SP2026', dry_run=True, populate_template=str(template_path), populate_output=str(out_path))

    assert out_path.exists()
    wb2 = load_workbook(str(out_path), data_only=True)
    main = wb2['SP2026']

    # Find the row with date 2026-01-19
    # Robust ISO date check
    def _iso_date_str(v):
        import datetime
        if v is None:
            return None
        if hasattr(v, 'date'):
            return v.date().isoformat()
        if isinstance(v, datetime.date):
            return v.isoformat()
        if isinstance(v, str):
            from scripts.import_calendar_xls import parse_any_date
            return parse_any_date(v) or v
        return None

    found_row = None
    for r in range(1, main.max_row + 1):
        v = main.cell(row=r, column=1).value
        if _iso_date_str(v) == '2026-01-19':
            found_row = r
            break
    assert found_row is not None, 'Holiday date not written to main sheet'

    # Assignment short (col 5) should be 'MLK'
    assert main.cell(row=found_row, column=5).value == 'MLK'

    # Row should be shaded (patternType / fill_type solid)
    assert main.cell(row=found_row, column=1).fill is not None
    assert getattr(main.cell(row=found_row, column=1).fill, 'fill_type', None) == 'solid' or getattr(main.cell(row=found_row, column=1).fill, 'patternType', None) == 'solid'