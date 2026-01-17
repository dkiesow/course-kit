import sqlite3
import tempfile
import os
from pathlib import Path

import pytest

pytest.importorskip('openpyxl')

from scripts.import_calendar_xls import parse_xls_row, import_calendar_xls
from openpyxl import Workbook


def test_parse_xls_row_basic():
    row = ("Unit 1 - 1/11/26", "Ch. 1", "Intro", "Discussion", "HW1, Lab1")
    parsed = parse_xls_row(row, semester='SP2026')
    assert parsed['unit'] is not None
    assert parsed['dates'] == ['2026-01-11']
    assert parsed['reading_list'] == 'Ch. 1'
    assert parsed['monday_details'] == 'Intro'
    assert parsed['wednesday_details'] == 'Discussion'
    assert parsed['assignments'] == ['HW1', 'Lab1']


def test_parse_xls_row_multiple_dates_no_year():
    row = ('Unit 1 - 1/26 - 1/28', '', '', '', '')
    parsed = parse_xls_row(row, semester='SP2026')
    assert parsed['dates'] == ['2026-01-26', '2026-01-28']


def test_parse_xls_row_handles_datetime_cell():
    import datetime as _dt
    row = (_dt.datetime(2026, 1, 21, 0, 0, 0), '', '', '', '')
    parsed = parse_xls_row(row)
    assert parsed['dates'] == ['2026-01-21']


def test_integration_update_deck(tmp_path):
    # Create a temporary DB with minimal schema
    db_path = tmp_path / "test_presentations.db"
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            order_index INTEGER,
            unit TEXT,
            reading_list TEXT,
            monday_details TEXT,
            wednesday_details TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT NOT NULL,
            short TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE deck_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            deck_id INTEGER NOT NULL,
            assignment_id INTEGER NOT NULL,
            UNIQUE(deck_id, assignment_id)
        )
    ''')
    # Insert a deck with date in ISO format
    c.execute('INSERT INTO decks (id, date) VALUES (?, ?)', (1, '2026-01-11'))
    # Insert an assignment with short code
    c.execute('INSERT INTO assignments (id, semester, name, due_date, short) VALUES (?, ?, ?, ?, ?)', (1, 'SP2026', 'Homework 1', '2026-01-15', 'HW1'))
    conn.commit()
    conn.close()

    # Create a simple workbook
    wb = Workbook()
    ws = wb.active
    ws.append(['Unit', 'Reading', 'MON', 'WED', 'Assignments'])
    ws.append(['Unit 1 - 1/11/26', 'Ch. 1', 'Intro', 'Discussion', 'HW1'])
    xls_path = tmp_path / 'cal.xlsx'
    wb.save(str(xls_path))

    # Run importer in dry-run with linking
    updated, linked, skipped = import_calendar_xls(str(xls_path), db_path=str(db_path), semester='SP2026', dry_run=True, link_assignments=True, apply_readings='first')
    assert updated == 1
    assert linked == 1
    assert skipped == 0

    # Run importer for real
    updated2, linked2, skipped2 = import_calendar_xls(str(xls_path), db_path=str(db_path), semester='SP2026', dry_run=False, link_assignments=True, apply_readings='first')
    assert updated2 == 1
    assert linked2 == 1
    assert skipped2 == 0

    # Verify DB updates
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('SELECT unit, reading_list, monday_details, wednesday_details FROM decks WHERE id = 1')
    r = c.fetchone()
    assert r[0] is not None
    assert r[1] == 'Ch. 1'
    assert r[2] == 'Intro'
    assert r[3] == 'Discussion'

    c.execute('SELECT assignment_id FROM deck_assignments WHERE deck_id = 1')
    links = c.fetchall()
    assert len(links) == 1
    conn.close()


def test_link_multiple_assignments_with_same_short(tmp_path):
    db_path = tmp_path / 'test_pres_multi.db'
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            order_index INTEGER,
            unit TEXT,
            reading_list TEXT,
            monday_details TEXT,
            wednesday_details TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT NOT NULL,
            short TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE deck_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            deck_id INTEGER NOT NULL,
            assignment_id INTEGER NOT NULL,
            UNIQUE(deck_id, assignment_id)
        )
    ''')
    # Insert deck and two assignments with same short
    c.execute('INSERT INTO decks (id, date) VALUES (?, ?)', (1, '2026-01-11'))
    c.execute('INSERT INTO assignments (id, semester, name, due_date, short) VALUES (?, ?, ?, ?, ?)', (1, 'SP2026', 'Guest Q Part A', '2026-01-15', 'Guest Q'))
    c.execute('INSERT INTO assignments (id, semester, name, due_date, short) VALUES (?, ?, ?, ?, ?)', (2, 'SP2026', 'Guest Q Part B', '2026-01-16', 'Guest Q'))
    conn.commit()
    conn.close()

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(['Unit', 'Reading', 'MON', 'WED', 'Assignments'])
    ws2.append(['Unit 1 - 1/11/26', '', '', '', 'Guest Q'])
    xls_path2 = tmp_path / 'cal2.xlsx'
    wb2.save(str(xls_path2))

    # Dry run should report 2 links
    updated, linked, skipped = import_calendar_xls(str(xls_path2), db_path=str(db_path), semester='SP2026', dry_run=True, link_assignments=True)
    assert linked == 2

    # Apply for real
    updated2, linked2, skipped2 = import_calendar_xls(str(xls_path2), db_path=str(db_path), semester='SP2026', dry_run=False, link_assignments=True)
    assert linked2 == 2

    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('SELECT assignment_id FROM deck_assignments WHERE deck_id = 1 ORDER BY assignment_id')
    links = [r[0] for r in c.fetchall()]
    conn.close()
    assert set(links) == {1, 2}


def test_case_insensitive_short_matching(tmp_path):
    db_path = tmp_path / 'test_pres_case.db'
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            order_index INTEGER,
            unit TEXT,
            reading_list TEXT,
            monday_details TEXT,
            wednesday_details TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT NOT NULL,
            short TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE deck_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            deck_id INTEGER NOT NULL,
            assignment_id INTEGER NOT NULL,
            UNIQUE(deck_id, assignment_id)
        )
    ''')
    # Insert deck and an assignment with mixed-case short
    c.execute('INSERT INTO decks (id, date) VALUES (?, ?)', (1, '2026-01-11'))
    c.execute('INSERT INTO assignments (id, semester, name, due_date, short) VALUES (?, ?, ?, ?, ?)', (1, 'SP2026', 'Report Proposal', '2026-02-13', 'Report Proposal'))
    conn.commit()
    conn.close()

    wb3 = Workbook()
    ws3 = wb3.active
    ws3.append(['Unit', 'Reading', 'MON', 'WED', 'Assignments'])
    # XLS contains uppercase short
    ws3.append(['Unit 1 - 1/11/26', '', '', '', 'REPORT PROPOSAL'])
    xls_path3 = tmp_path / 'cal3.xlsx'
    wb3.save(str(xls_path3))

    updated, linked, skipped = import_calendar_xls(str(xls_path3), db_path=str(db_path), semester='SP2026', dry_run=True, link_assignments=True)
    assert linked == 1

    updated2, linked2, skipped2 = import_calendar_xls(str(xls_path3), db_path=str(db_path), semester='SP2026', dry_run=False, link_assignments=True)
    assert linked2 == 1

    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('SELECT assignment_id FROM deck_assignments WHERE deck_id = 1 ORDER BY assignment_id')
    links = [r[0] for r in c.fetchall()]
    conn.close()
    assert set(links) == {1}
