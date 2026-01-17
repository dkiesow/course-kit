import sqlite3
import pytest
pytest.importorskip('openpyxl')
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))
from pathlib import Path
from openpyxl import load_workbook, Workbook

from scripts.import_calendar_xls import import_calendar_xls


def test_populate_template_writes_xlsx(tmp_path):
    db_path = tmp_path / "test_template.db"
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT,
            short TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            order_index INTEGER,
            unit TEXT,
            reading_list TEXT,
            monday_details TEXT,
            wednesday_details TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE deck_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            deck_id INTEGER NOT NULL,
            assignment_id INTEGER NOT NULL,
            UNIQUE(deck_id, assignment_id)
        )
    ''')
    # Insert a few assignments for SP2026
    c.execute('INSERT INTO assignments (id, semester, name, due_date, short) VALUES (?, ?, ?, ?, ?)',
              (1, 'SP2026', 'Report Proposal', '2026-02-13', 'Report Proposal'))
    c.execute('INSERT INTO assignments (id, semester, name, due_date, short) VALUES (?, ?, ?, ?, ?)',
              (2, 'SP2026', 'Unit One Quiz', '2026-02-13', 'Quiz One'))
    conn.commit()
    conn.close()

    # Create a minimal template workbook with a Dates sheet
    template_path = tmp_path / 'template.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'SP2026'
    wb.create_sheet('SP2026 Dates')
    wb.save(str(template_path))

    out_path = tmp_path / 'populated.xlsx'

    # Run importer which should populate the template and write out_path
    import_calendar_xls(str(template_path), db_path=str(db_path), semester='SP2026', dry_run=True, populate_template=str(template_path), populate_output=str(out_path))

    assert out_path.exists()
    wb2 = load_workbook(str(out_path))
    assert 'SP2026 Dates' in wb2.sheetnames
    ws2 = wb2['SP2026 Dates']
    # Check that at least the first two rows contain the inserted assignments (column B = name, C = due date, D = short)
    assert ws2.cell(row=1, column=2).value == 'Report Proposal'
    assert ws2.cell(row=2, column=2).value == 'Unit One Quiz'
    # due date should be a date object for at least the first row
    assert ws2.cell(row=1, column=3).value is not None


def test_populate_template_writes_topics(tmp_path):
    db_path = tmp_path / "test_template_topics.db"
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            topic1 TEXT,
            topic2 TEXT
        )
    ''')
    # Minimal assignments table to satisfy finder (may be empty for this test)
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT,
            short TEXT
        )
    ''')
    c.execute('INSERT INTO decks (id, presentation_id, week, date, topic1, topic2) VALUES (?, ?, ?, ?, ?, ?)',
              (1, 1, 'Week 1', '2026-01-13', 'Topic Alpha', 'Topic Beta'))
    conn.commit()
    conn.close()

    # Create template with main sheet placeholders including TOPIC1/TOPIC2
    template_path = tmp_path / 'template_topics.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'SP2026'
    # Date token in column A (parsed as 1/13 -> 2026-01-13) and placeholders for topics
    ws.append(['1/13', '{READINGS}', '{MON}', '{WED}', '{ASSIGNMENT}', '{TOPIC1}', '{TOPIC2}'])
    wb.create_sheet('SP2026 Dates')
    wb.save(str(template_path))

    out_path = tmp_path / 'populated_topics.xlsx'

    import_calendar_xls(str(template_path), db_path=str(db_path), semester='SP2026', dry_run=True, populate_template=str(template_path), populate_output=str(out_path))

    assert out_path.exists()
    wb2 = load_workbook(str(out_path))
    ws2 = wb2['SP2026']
    # TOPIC1 -> column 6, TOPIC2 -> column 7 in the first row
    assert ws2.cell(row=1, column=6).value == 'Topic Alpha'
    assert ws2.cell(row=1, column=7).value == 'Topic Beta'
