import sqlite3
import pytest
pytest.importorskip('openpyxl')
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))
from pathlib import Path
from openpyxl import load_workbook, Workbook

from scripts.import_calendar_xls import import_calendar_xls


def test_populate_template_block_layout(tmp_path):
    db_path = tmp_path / "test_template_layout.db"
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT,
            short TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            unit TEXT,
            reading_list TEXT,
            topic1 TEXT,
            topic2 TEXT,
            monday_details TEXT,
            wednesday_details TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE deck_assignments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            deck_id INTEGER NOT NULL,
            assignment_id INTEGER NOT NULL,
            UNIQUE(deck_id, assignment_id)
        )
    ''')

    # Insert deck with topics, readings, and two assignments
    c.execute('INSERT INTO decks (id, presentation_id, week, date, unit, reading_list, topic1, topic2, monday_details, wednesday_details) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)',
              (1, 7, 'Week 1', '2026-01-21', None, 'Read1\nRead2\nRead3\nRead4', 'MonTopic1', 'MonTopic2', 'MonDetail', 'WedDetail'))
    c.execute('INSERT INTO assignments (id, semester, name, due_date, short) VALUES (?, ?, ?, ?, ?)', (1, 'SP2026', 'Homework1', '2026-01-25', 'HW1'))
    c.execute('INSERT INTO assignments (id, semester, name, due_date, short) VALUES (?, ?, ?, ?, ?)', (2, 'SP2026', 'Homework2', '2026-02-01', 'HW2'))
    c.execute('INSERT INTO deck_assignments (deck_id, assignment_id) VALUES (?, ?)', (1, 1))
    c.execute('INSERT INTO deck_assignments (deck_id, assignment_id) VALUES (?, ?)', (1, 2))
    conn.commit()
    conn.close()

    # Create a template with a 3-row block for the week
    template_path = tmp_path / 'template_block.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'SP2026'
    # Top row: Week | Readings | Mon Topic | Wed Topic | Assign
    ws.append(['{WEEK #}', '{readings}', 'Topic', 'Topic', '{ASSIGNMENT}'])
    # Second row: Dates | Readings | Topic | Topic | {ASSIGNMENT}
    ws.append(['{DATES}', '{readings}', 'Topic', 'Topic', '{ASSIGNMENT}'])
    # Third content row for extra lines
    ws.append([None, '{readings}', 'Topic', 'Topic', '{ASSIGNMENT}'])
    wb.create_sheet('SP2026 Dates')
    wb.save(str(template_path))

    out_path = tmp_path / 'populated_block.xlsx'
    import_calendar_xls(str(template_path), db_path=str(db_path), semester='SP2026', dry_run=True, populate_template=str(template_path), populate_output=str(out_path), populate_from_db=True, presentation_id=7)

    assert out_path.exists()
    wb2 = load_workbook(str(out_path), data_only=True)
    main = wb2['SP2026']

    # Week label should be 'One' (words)
    assert main.cell(row=1, column=1).value == 'One'
    # Date should be formatted as M/D
    assert main.cell(row=2, column=1).value == '1/21'

    # Readings split into rows 1..3 of column 2
    assert main.cell(row=1, column=2).value == 'Read1'
    assert main.cell(row=2, column=2).value == 'Read2'
    assert main.cell(row=3, column=2).value == 'Read3'

    # Mon topics (col 3)
    assert main.cell(row=1, column=3).value == 'MonTopic1'
    assert main.cell(row=2, column=3).value == 'MonTopic2'
    assert main.cell(row=3, column=3).value == 'MonDetail'

    # Wed topics (col 4) include same topics + wed detail as third row
    assert main.cell(row=1, column=4).value == 'MonTopic1'
    assert main.cell(row=2, column=4).value == 'MonTopic2'
    assert main.cell(row=3, column=4).value == 'WedDetail'

    # Assignments appear in rows 2 and 3 of column 5
    assert main.cell(row=2, column=5).value == 'HW1'
    assert main.cell(row=3, column=5).value == 'HW2'
