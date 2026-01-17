import sqlite3
import pytest
pytest.importorskip('openpyxl')
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))
from openpyxl import load_workbook, Workbook
from scripts.import_calendar_xls import import_calendar_xls


def test_populate_template_with_literal_topic_cells(tmp_path):
    db_path = tmp_path / "test_template_topic_literal.db"
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            topic1 TEXT,
            topic2 TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT,
            short TEXT
        )
    ''')
    c.execute('INSERT INTO decks (id, presentation_id, week, date, topic1, topic2) VALUES (?, ?, ?, ?, ?, ?)',
              (1, 7, 'Week 1', '2026-01-21', 'Literal A', 'Literal B'))
    conn.commit()
    conn.close()

    template_path = tmp_path / 'template_topic_literal.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'SP2026'
    # header + block that uses literal 'Topic' cells
    ws.append(['Header', '', '', '', '', '', '', ''])
    ws.append(['{UNIT #}', None, None, None, None, None, None, None])
    ws.append(['{WEEK #}', '{readings}', 'Topic', 'Topic', '{ASSIGNMENT}', None, None, None])
    ws.append(['{DATES}', '{readings}', 'Topic', 'Topic', '{ASSIGNMENT}', None, None, None])
    wb.create_sheet('SP2026 Dates')
    wb.save(str(template_path))

    out_path = tmp_path / 'pop_topic_literal.xlsx'

    import_calendar_xls(str(template_path), db_path=str(db_path), semester='SP2026', dry_run=True, populate_template=str(template_path), populate_output=str(out_path), populate_from_db=True, presentation_id=7)

    assert out_path.exists()
    wb2 = load_workbook(str(out_path), data_only=True)
    main = wb2['SP2026']
    # The 'Topic' cells (col3, col4) in the first populated block should show the DB topics
    assert main.cell(row=4, column=3).value == 'Literal A'
    assert main.cell(row=4, column=4).value == 'Literal B'