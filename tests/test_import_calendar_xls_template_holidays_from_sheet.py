import sqlite3
from pathlib import Path
from openpyxl import Workbook, load_workbook
import pytest

pytest.importorskip('openpyxl')
import sys, pathlib
sys.path.insert(0, str(pathlib.Path(__file__).resolve().parents[1]))

from scripts.import_calendar_xls import import_calendar_xls


def test_holidays_from_data_sheet(tmp_path):
    # Minimal DB
    db_path = tmp_path / 'test_holidays_sheet.db'
    conn = sqlite3.connect(str(db_path))
    c = conn.cursor()
    c.execute('''
        CREATE TABLE assignments (
            id INTEGER PRIMARY KEY,
            semester TEXT,
            name TEXT NOT NULL,
            due_date TEXT,
            short TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE decks (
            id INTEGER PRIMARY KEY,
            presentation_id INTEGER,
            week TEXT,
            date TEXT,
            order_index INTEGER,
            unit TEXT,
            reading_list TEXT,
            monday_details TEXT,
            wednesday_details TEXT
        )
    ''')
    conn.commit()
    conn.close()

    # Template workbook with SP2026 Dates containing holiday entries
    template_path = tmp_path / 'template_holidays_sheet.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = 'SP2026'
    ws['A10'] = '{DATES}'
    ws['B10'] = '{readings}'
    # Dates data sheet
    sds = wb.create_sheet('SP2026 Dates')
    # Row format: Semester, Week, Name, Date, Short
    sds.append(['SP2026', 1, 'Holiday', '1/19/26', 'MLK'])
    sds.append(['SP2026', 13, 'No Class', '4/15/26', 'No Class'])
    wb.save(str(template_path))

    # Calendar workbook with rows for the dates
    cal_path = tmp_path / 'calendar_sheet.xlsx'
    cw = Workbook()
    ws2 = cw.active
    ws2.append(['Unit', 'Reading', 'MON', 'WED', 'Assignments'])
    ws2.append(['Unit 1 - 1/19/26', 'Ch. 1', '', '', ''])
    ws2.append(['Unit 13 - 4/15/26', 'Ch. 13', '', '', ''])
    cw.save(str(cal_path))

    out_path = tmp_path / 'populated_holidays_sheet.xlsx'

    import_calendar_xls(str(cal_path), db_path=str(db_path), semester='SP2026', dry_run=True, populate_template=str(template_path), populate_output=str(out_path))

    assert out_path.exists()
    wb2 = load_workbook(str(out_path), data_only=True)
    main = wb2['SP2026']

    # Check MLK row
    def _iso_date_str(v):
        import datetime
        if v is None:
            return None
        if hasattr(v, 'date'):
            return v.date().isoformat()
        if isinstance(v, datetime.date):
            return v.isoformat()
        from scripts.import_calendar_xls import parse_any_date
        return parse_any_date(v) or v

    found_mlk = None
    found_no_class = None
    for r in range(1, main.max_row + 1):
        v = main.cell(row=r, column=1).value
        if _iso_date_str(v) == '2026-01-19':
            found_mlk = r
        if _iso_date_str(v) == '2026-04-15':
            found_no_class = r
    assert found_mlk is not None, 'MLK date not written'
    assert found_no_class is not None, 'No Class date not written'

    # Column E should have short labels
    assert main.cell(row=found_mlk, column=5).value == 'MLK'
    assert main.cell(row=found_no_class, column=5).value == 'No Class'

    # Row shading
    assert getattr(main.cell(row=found_mlk, column=1).fill, 'fill_type', None) == 'solid' or getattr(main.cell(row=found_mlk, column=1).fill, 'patternType', None) == 'solid'
    assert getattr(main.cell(row=found_no_class, column=1).fill, 'fill_type', None) == 'solid' or getattr(main.cell(row=found_no_class, column=1).fill, 'patternType', None) == 'solid'
