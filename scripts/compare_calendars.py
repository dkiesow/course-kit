#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Alignment

# Load both workbooks
exported = openpyxl.load_workbook('/private/tmp/SP2026_calendar.xlsx')
original = openpyxl.load_workbook('/Users/kiesowd/Library/Mobile Documents/com~apple~CloudDocs/Documents/Mizzou/Spring 2026/J4734/4734-master.xlsx')

ws_exported = exported.active
ws_original = original.active

print("=" * 80)
print("CALENDAR COMPARISON: Exported vs Original")
print("=" * 80)

# 1. STRUCTURE
print("\n1. STRUCTURE:")
print(f"   Exported: {ws_exported.max_row} rows x {ws_exported.max_column} cols")
print(f"   Original: {ws_original.max_row} rows x {ws_original.max_column} cols")

# 2. CONTENT SAMPLE
print("\n2. CONTENT SAMPLE (First 15 rows, columns A-E):")
for row_num in range(1, min(16, ws_exported.max_row + 1)):
    exp_row = [ws_exported[f'{col}{row_num}'].value for col in ['A', 'B', 'C', 'D', 'E']]
    orig_row = [ws_original[f'{col}{row_num}'].value for col in ['A', 'B', 'C', 'D', 'E']]
    
    exp_str = " | ".join([str(v)[:15] if v else "" for v in exp_row])
    orig_str = " | ".join([str(v)[:15] if v else "" for v in orig_row])
    
    match = "✓" if exp_row == orig_row else "❌"
    print(f"\n   Row {row_num:2d} {match}")
    print(f"   Exp: {exp_str}")
    print(f"   Org: {orig_str}")

# 3. FONTS
print("\n3. FONT STYLES (Row 2, Col A):")
exp_cell = ws_exported['A2']
orig_cell = ws_original['A2']
print(f"   Exported: name={exp_cell.font.name}, size={exp_cell.font.size}, bold={exp_cell.font.bold}")
print(f"   Original: name={orig_cell.font.name}, size={orig_cell.font.size}, bold={orig_cell.font.bold}")

# 4. COLORS
print("\n4. FILL COLORS (Row 2, Col A):")
exp_fill = exp_cell.fill.start_color.rgb if exp_cell.fill.start_color else None
orig_fill = orig_cell.fill.start_color.rgb if orig_cell.fill.start_color else None
print(f"   Exported: {exp_fill}")
print(f"   Original: {orig_fill}")

# 5. BORDERS
print("\n5. BORDERS (Row 2, Col A):")
print(f"   Exported: top={exp_cell.border.top.style if exp_cell.border.top else None}")
print(f"   Original: top={orig_cell.border.top.style if orig_cell.border.top else None}")

# 6. COLUMN WIDTHS
print("\n6. COLUMN WIDTHS:")
for col in ['A', 'B', 'C', 'D', 'E']:
    exp_w = ws_exported.column_dimensions[col].width
    orig_w = ws_original.column_dimensions[col].width
    match = "✓" if abs(exp_w - orig_w) < 0.5 else "❌"
    print(f"   {col}: Exp={exp_w:.1f}, Orig={orig_w:.1f} {match}")

# 7. ROW HEIGHTS (sample)
print("\n7. ROW HEIGHTS (sample first 20):")
height_diffs = []
for i in range(1, 21):
    exp_h = ws_exported.row_dimensions[i].height
    orig_h = ws_original.row_dimensions[i].height
    if exp_h != orig_h:
        height_diffs.append((i, exp_h, orig_h))

if height_diffs:
    for row, exp_h, orig_h in height_diffs[:10]:
        print(f"   Row {row:2d}: Exp={exp_h}, Orig={orig_h}")
else:
    print("   All heights match ✓")

# 8. KEY MARKERS
print("\n8. KEY MARKERS:")
unit_exp = sum(1 for row in ws_exported.iter_rows() if any(c.value and 'UNIT' in str(c.value) for c in row))
unit_orig = sum(1 for row in ws_original.iter_rows() if any(c.value and 'UNIT' in str(c.value) for c in row))
print(f"   UNIT headers: Exp={unit_exp}, Orig={unit_orig}")

print("\n" + "=" * 80)
