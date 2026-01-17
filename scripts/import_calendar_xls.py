#!/usr/bin/env python3
"""Import calendar rows from an Excel template into the decks table and link assignments.

Expected columns (first sheet by default):
A: Unit / Date(s)
B: Reading list
C: MON details
D: WED details
E: Assignments (comma-separated list of short codes or names)

Behavior:
- Parse dates from column A and match decks by date (attempts multiple common date formats).
- Update `decks` fields: `unit`, `reading_list`, `monday_details`, `wednesday_details`.
- Optionally link assignments by `short` code using `--link-assignments` (requires `--semester`).
- Supports `--dry-run` and `--backup` behaviors similar to `scripts/import_assignments.py`.

Note: This script prefers not to create new assignments. It will warn when assignment short codes are not found.
"""

from __future__ import annotations

import argparse
import csv
import datetime
import json
import os
import re
import shutil
import sqlite3
import sys
from typing import List, Dict, Optional, Tuple, Set
from copy import copy

# Import parse_date from the assignments importer to keep consistent date normalization
try:
    from scripts.import_assignments import parse_date
except Exception:
    # Fallback parse_date defined locally if the import fails
    def parse_date(s: str) -> Optional[str]:
        if not s:
            return None
        s = s.strip()
        m = re.search(r"(\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4})", s)
        if not m:
            return None
        tok = m.group(1).replace('.', '/').replace('-', '/')
        for fmt in ("%m/%d/%y", "%m/%d/%Y"):
            try:
                dt = datetime.datetime.strptime(tok, fmt).date()
                return dt.isoformat()
            except ValueError:
                continue
        return None


def extract_iso_dates(value: object, semester: Optional[str] = None) -> List[str]:
    """Extract one or more ISO (YYYY-MM-DD) date strings from various cell contents.

    Handles:
    - datetime.date and datetime.datetime objects
    - ISO strings with or without time (e.g., '2026-01-21' or '2026-01-21 00:00:00')
    - mm/dd or mm/dd/yy or mm/dd/yyyy tokens; if year is missing, infer from `semester` (e.g., 'SP2026') or fall back to current year
    - ranges like '1/26 - 1/28' (returns both dates)
    """
    if value is None:
        return []
    # direct date/datetime objects
    if isinstance(value, datetime.datetime):
        return [value.date().isoformat()]
    if isinstance(value, datetime.date):
        return [value.isoformat()]

    s = str(value).strip()
    if not s:
        return []


    # ISO-like values (with or without time)
    try:
        d = datetime.date.fromisoformat(s)
        return [d.isoformat()]
    except Exception:
        try:
            dt = datetime.datetime.fromisoformat(s)
            return [dt.date().isoformat()]
        except Exception:
            pass

    # Find mm/dd tokens with optional year
    tokens = re.findall(r"(\d{1,2}[\/\-.]\d{1,2}(?:[\/\-.]\d{2,4})?)", s)
    if not tokens:
        return []

    # Determine default year from semester if possible
    default_year = None
    if semester:
        m = re.search(r"(\d{4})", semester)
        if m:
            default_year = int(m.group(1))
        else:
            m2 = re.search(r"(\d{2})$", semester)
            if m2:
                default_year = 2000 + int(m2.group(1))
    if default_year is None:
        default_year = datetime.date.today().year

    dates: List[str] = []
    for tok in tokens:
        tok = tok.strip()
        # If token includes a YEAR segment (three-part mm/dd/yy or mm/dd/yyyy)
        if re.search(r"\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4}$", tok):
            try:
                d_iso = parse_date(tok)
            except Exception:
                d_iso = None
            if d_iso and d_iso not in dates:
                dates.append(d_iso)
            continue
        # No year => append default year and parse
        try:
            d_iso = parse_date(f"{tok}/{default_year}")
            if d_iso and d_iso not in dates:
                dates.append(d_iso)
        except Exception:
            continue
    return dates


def parse_xls_row(values: Tuple, semester: Optional[str] = None) -> Dict[str, Optional[object]]:
    """Parse a row's values (A..E) into a structured dict.

    values: tuple-like from `openpyxl` row `values_only=True` or any sequence of cell values.
    Returns: dict with keys: unit (str|None), dates (list[str]), reading_list (str|None), monday_details (str|None), wednesday_details (str|None), assignments (list[str])
    """
    a = values[0] if len(values) >= 1 else None
    b = values[1] if len(values) >= 2 else None
    c = values[2] if len(values) >= 3 else None
    d = values[3] if len(values) >= 4 else None
    e = values[4] if len(values) >= 5 else None

    unit = None
    if a:
        ma = re.search(r"Unit\s*#?\s*(\d+)", str(a), flags=re.I)
        if ma:
            unit = ma.group(1)
        else:
            # Try to extract plain leading token
            s = str(a).strip()
            if s:
                unit = s

    # Extract dates (handles ranges and no-year mm/dd tokens)
    dates: List[str] = extract_iso_dates(a, semester=semester)

    reading_list = str(b).strip() if b not in (None, "") else None
    monday_details = str(c).strip() if c not in (None, "") else None
    wednesday_details = str(d).strip() if d not in (None, "") else None

    assignments = []
    if e:
        parts = re.split(r"[,;|]+", str(e))
        assignments = [p.strip() for p in parts if p and p.strip()]

    return {
        "unit": unit,
        "dates": dates,
        "reading_list": reading_list,
        "monday_details": monday_details,
        "wednesday_details": wednesday_details,
        "assignments": assignments,
    }


def parse_any_date(s: Optional[str]) -> Optional[str]:
    """Try to parse various date formats and return ISO date string or None."""
    if not s:
        return None
    s = s.strip()
    # Iso format
    try:
        return datetime.date.fromisoformat(s).isoformat()
    except Exception:
        pass
    # Token search + parse_date
    m = re.search(r"(\d{1,2}[\/\-.]\d{1,2}[\/\-.]\d{2,4})", s)
    if m:
        return parse_date(m.group(1))
    return None


def find_deck_ids_by_date(conn: sqlite3.Connection, iso_date: str) -> List[int]:
    c = conn.cursor()
    c.execute("SELECT id, date FROM decks WHERE date IS NOT NULL")
    res = []
    for row_id, row_date in c.fetchall():
        parsed = parse_any_date(row_date)
        if parsed and parsed == iso_date:
            res.append(row_id)
    return res


# Helper: convert small integers to English words for display (1 -> 'One', 2 -> 'Two', ...)
def num_to_words(n: int) -> str:
    ones = ["Zero","One","Two","Three","Four","Five","Six","Seven","Eight","Nine","Ten","Eleven","Twelve","Thirteen","Fourteen","Fifteen","Sixteen","Seventeen","Eighteen","Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    try:
        ni = int(n)
    except Exception:
        return str(n)
    if ni < 20:
        return ones[ni]
    if ni < 100:
        t = ni // 10
        r = ni % 10
        if r == 0:
            return tens[t]
        return f"{tens[t]}-{ones[r]}"
    return str(n)


def _norm_identifier(s: Optional[str]) -> str:
    if s is None:
        return ''
    # Collapse whitespace and uppercase for deterministic compares
    return re.sub(r"\s+", ' ', str(s).strip()).upper()


def find_assignment_candidates(conn: sqlite3.Connection, assignment_identifier: str, semester: str) -> List[Tuple[int, str]]:
    """Return a list of (id, name) candidates that match `assignment_identifier` by short (preferred) or name.

    Matching is case-insensitive and whitespace-normalized.
    """
    c = conn.cursor()
    ident_norm = _norm_identifier(assignment_identifier)
    candidates: List[Tuple[int, str]] = []

    # Look for short matches first
    c.execute('SELECT id, short, name FROM assignments WHERE semester = ? AND short IS NOT NULL', (semester,))
    for aid, short, name in c.fetchall():
        if _norm_identifier(short) == ident_norm:
            candidates.append((aid, name))
    if candidates:
        return candidates

    # Fall back to name matches
    c.execute('SELECT id, name FROM assignments WHERE semester = ? AND name IS NOT NULL', (semester,))
    for aid, name in c.fetchall():
        if _norm_identifier(name) == ident_norm:
            candidates.append((aid, name))
    return candidates


def link_assignment_to_deck(conn: sqlite3.Connection, deck_id: int, assignment_identifier: str, semester: str) -> List[int]:
    """Find assignments by `short` or `name` and link all matches to deck via deck_assignments. Returns list of linked assignment ids.

    Matching is case-insensitive and whitespace-normalized to accommodate XLS that renders short codes in ALL CAPS.
    """
    c = conn.cursor()
    candidates = find_assignment_candidates(conn, assignment_identifier, semester)
    linked_ids: List[int] = []
    for aid, _name in candidates:
        c.execute('INSERT OR IGNORE INTO deck_assignments (deck_id, assignment_id) VALUES (?, ?)', (deck_id, aid))
        linked_ids.append(aid)
    if linked_ids:
        conn.commit()
    return linked_ids


def import_calendar_xls(
    xls_path: str,
    db_path: str = 'presentations.db',
    semester: Optional[str] = None,
    backup: bool = False,
    dry_run: bool = True,
    link_assignments: bool = False,
    create_decks: bool = False,
    report_csv: Optional[str] = None,
    apply_readings: str = 'none',
    link_by_due: bool = False,
    link_by_due_n: int = 3,
    export_template_csv: Optional[str] = None,
    populate_template: Optional[str] = None,
    populate_output: Optional[str] = None,
    populate_base: Optional[str] = None,
    no_class_dates: Optional[Dict[str, str]] = None,
    no_class_csv: Optional[str] = None,
    # When populating templates, use DB decks only (ignore source workbook rows)
    populate_from_db: bool = False,
    presentation_id: Optional[int] = None,
    # Clear all semester data before importing (assignments, links, readings)
    clear_semester_data: bool = False,
) -> Tuple[int, int, int]:
    """Import calendar rows. Returns (updated_decks, linked_assignments, skipped_rows).

    If `report_csv` is provided, write a CSV describing proposed updates and predicted assignment links. Recommended to use with `--dry-run`.

    `apply_readings` controls whether the `reading_list` from the XLS is written into deck records:
      - 'none' (default): do not modify deck `reading_list` fields (reading lists kept for XLS only and included in report)
      - 'first': write reading_list to the first deck matched for that row
      - 'all': write reading_list to all matched decks
      - 'monday': prefer writing to the deck whose date is a Monday, or fallback to first
      - 'per_week': aggregate reading entries for a week and write once to a canonical deck for that week

    `link_by_due` (opt-in) will link existing assignments (in `assignments` table) to the N prior decks before each assignment's due date (default N=3). This mirrors the manual workflow of inserting upcoming assignment details into the preceding decks.
    """
    print(f"[DEBUG] import_calendar_xls called with xls_path={xls_path}, db_path={db_path}, semester={semester}, dry_run={dry_run}, apply_readings={apply_readings}")
    try:
        from openpyxl import load_workbook
    except Exception as e:
        raise RuntimeError("The 'openpyxl' package is required. Install with 'pip install openpyxl'.") from e

    if backup and os.path.exists(db_path):
        ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        bak_path = f"{db_path}.bak.{ts}"
        shutil.copy2(db_path, bak_path)
        print(f"Backup created: {bak_path}")

    # Load workbook or CSV
    if xls_path.lower().endswith('.csv'):
        import csv
        with open(xls_path, 'r', newline='', encoding='utf-8-sig', errors='replace') as f:
            # CSV is actually comma-delimited, need to handle properly
            content = f.read()
            # Check if it's comma or tab delimited
            lines = content.split('\n')
            if lines and '\t' in lines[0] and ',' in lines[0]:
                # Has both, need to determine which is the delimiter
                # Count commas vs tabs in first line
                comma_count = lines[0].count(',')
                tab_count = lines[0].count('\t')
                delimiter = '\t' if tab_count > comma_count else ','
            elif lines and '\t' in lines[0]:
                delimiter = '\t'
            else:
                delimiter = ','
            
            f.seek(0)
            reader = csv.reader(f, delimiter=delimiter)
            csv_rows = list(reader)
        wb = None
        ws = None
    else:
        wb = load_workbook(xls_path, data_only=True)
        print('[DEBUG] workbook loaded')
        ws = wb.active
        print('[DEBUG] worksheet active')
        csv_rows = None

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    # If requested, clear all semester data before importing to make XLS authoritative
    if clear_semester_data and semester:
        print(f"[DEBUG] Clearing all assignments, links, and readings for semester {semester}")
        if not dry_run:
            # Delete links first to avoid orphaned links
            cursor.execute('DELETE FROM deck_assignments WHERE assignment_id IN (SELECT id FROM assignments WHERE semester = ?)', (semester,))
            # Delete assignments for the semester
            cursor.execute('DELETE FROM assignments WHERE semester = ?', (semester,))
            # Clear reading_list for decks (assuming decks are linked to semester via presentation, but to clear all for semester)
            # Actually, decks don't have semester, but since we're importing for semester, perhaps clear readings for all decks if apply_readings
            # But for now, since apply_readings will set them, we can clear reading_list for all decks
            cursor.execute('UPDATE decks SET reading_list = NULL')
            conn.commit()
        else:
            print(f"[DRY RUN] Would clear assignments, links, and readings for semester {semester}")

    # Normalize and load no-class dates mapping (from parameter or CSV)
    no_class_map: Dict[str, str] = {}
    if no_class_dates:
        for k, v in no_class_dates.items():
            if not k:
                continue
            parsed = parse_any_date(k) or k
            if parsed:
                no_class_map[parsed] = v or ''
    if no_class_csv:
        if os.path.exists(no_class_csv):
            with open(no_class_csv, 'r', newline='') as nf:
                r = csv.reader(nf)
                for row in r:
                    if not row:
                        continue
                    date_tok = row[0]
                    label = row[1] if len(row) > 1 else ''
                    parsed = parse_any_date(date_tok) or date_tok
                    if parsed:
                        no_class_map[parsed] = label or ''
            print(f"Loaded {len(no_class_map)} no-class dates from {no_class_csv}")
        else:
            print(f"No-class CSV not found: {no_class_csv}")

    def _link_assignments_by_due(sem: str, n_prior: int) -> int:
        """Link assignments with due dates (semester `sem`) to the N prior decks before each due_date.

        Returns number of links created (or predicted in dry_run). Adds corresponding planned_actions entries.
        """
        created = 0
        if not sem:
            return 0
        c = conn.cursor()
        # Fetch assignments with due dates for semester
        c.execute('SELECT id, name, due_date, short FROM assignments WHERE semester = ? AND due_date IS NOT NULL', (sem,))
        rows = c.fetchall()
        if not rows:
            return 0
        # Build deck list with parsed dates
        c.execute("SELECT id, date FROM decks WHERE date IS NOT NULL")
        deck_rows = []
        for did, dtext in c.fetchall():
            parsed = parse_any_date(dtext)
            if parsed:
                deck_rows.append((did, parsed))
        # Sort decks by date ascending
        deck_rows.sort(key=lambda x: x[1])

        for aid, aname, adue, ashort in rows:
            try:
                due_iso = adue if isinstance(adue, str) and adue.strip() else None
                if not due_iso:
                    continue
                # Ensure normalized ISO
                due_parsed = parse_any_date(due_iso) or due_iso
                due_obj = datetime.date.fromisoformat(due_parsed)
            except Exception:
                continue

            # Find decks with date strictly before due_date
            candidates = [(did, d_iso) for did, d_iso in deck_rows if datetime.date.fromisoformat(d_iso) < due_obj]
            if not candidates:
                continue
            # pick most recent N prior decks
            candidates.sort(key=lambda x: x[1], reverse=True)
            chosen = candidates[:n_prior]
            for did, d_iso in chosen:
                # Format assignment display (human readable due date)
                due_norm = parse_any_date(adue) or adue
                try:
                    due_obj = datetime.date.fromisoformat(due_norm)
                    due_str = due_obj.strftime('%b %d, %Y')
                except Exception:
                    due_str = due_norm
                assign_display = f"{aname} — {due_str}"

                # Try to merge into an existing planned_action for this deck so the report shows assignments inline
                merged = False
                for act in planned_actions:
                    if act.get('deck_id') == did:
                        # Append to act['updates']['upcoming_assignments'] without duplication
                        upd = act.setdefault('updates', {})
                        cur = upd.get('upcoming_assignments', '')
                        cur_lines = [s.strip() for s in cur.split('\n') if s.strip()] if cur else []
                        if assign_display not in cur_lines:
                            cur_lines.append(assign_display)
                            upd['upcoming_assignments'] = '\n'.join(cur_lines)
                        # Also record as predicted assignment for visibility
                        atoks = act.setdefault('assignment_tokens', [])
                        token = (ashort or aname)
                        if token and token not in atoks:
                            atoks.append(token)
                        pids = act.setdefault('predicted_assignment_ids', [])
                        if aid not in pids:
                            pids.append(aid)
                        pnames = act.setdefault('predicted_assignment_names', [])
                        if aname not in pnames:
                            pnames.append(aname)
                        # record predicted due date (ISO) for visibility
                        pdues = act.setdefault('predicted_assignment_dues', [])
                        due_iso_norm = parse_any_date(adue) or (adue if adue else '')
                        if due_iso_norm and due_iso_norm not in pdues:
                            pdues.append(due_iso_norm)
                        merged = True
                        break

                if not merged:
                    # Add planned action for report when no deck action exists
                    due_iso_norm = parse_any_date(adue) or (adue if adue else '')
                    planned_actions.append({
                        'row': f'link_by_due_{aid}',
                        'first': '',
                        'date_iso': d_iso,
                        'deck_id': did,
                        'updates': {'upcoming_assignments': assign_display},
                        'assignment_tokens': [ashort or aname],
                        'predicted_assignment_ids': [aid],
                        'predicted_assignment_names': [aname],
                        'predicted_assignment_dues': [due_iso_norm],
                    })

                if not dry_run:
                    c.execute('INSERT OR IGNORE INTO deck_assignments (deck_id, assignment_id) VALUES (?, ?)', (did, aid))
                    conn.commit()
                created += 1
        return created

    updated_decks = 0
    linked_assignments = 0
    skipped = 0
    updated_deck_ids: Set[int] = set()

    planned_actions: List[Dict] = []
    skipped_rows: List[Dict] = []

    # For exports that populate the XLS template
    parsed_rows: List[Dict] = []

    # Reading aggregation structures for different apply_readings modes
    deck_readings: Dict[int, List[str]] = {}
    weekly_readings: Dict[Tuple[int, int], Dict] = {}

    # If requested, build parsed_rows from the DB (topics from decks, calendar from CSV)
    if populate_from_db and (populate_template or populate_base):
        print(f"[DEBUG] Building parsed_rows: calendar from CSV, topics from DB (presentation_id={presentation_id})")
        
        # First read CSV for calendar structure, holidays, and assignments
        if not csv_rows:
            print("[ERROR] CSV file required for calendar structure when using --populate-from-db")
            return
        
        # Then get deck topics from DB
        q = 'SELECT id, week, unit, date, reading_list, topic1, topic2, monday_details, wednesday_details FROM decks WHERE date IS NOT NULL'
        params: Tuple = ()
        if presentation_id:
            q += ' AND presentation_id = ?'
            params = (presentation_id,)
        q += ' ORDER BY date'
        try:
            cursor.execute(q, params)
            deck_rows = cursor.fetchall()
        except sqlite3.OperationalError:
            # Fallback for minimal test DBs
            q2 = 'SELECT id, week, date FROM decks WHERE date IS NOT NULL'
            params2: Tuple = ()
            if presentation_id:
                q2 += ' AND presentation_id = ?'
                params2 = (presentation_id,)
            q2 += ' ORDER BY date'
            cursor.execute(q2, params2)
            deck_rows = [(r[0], r[1], None, r[2], None, None, None, None, None) for r in cursor.fetchall()]
        
        # Group decks by extracting week number from week name for topic overlay
        import re
        deck_topics_by_week = {}
        
        # Word to number mapping for week names
        week_words = {
            'one': 1, 'two': 2, 'three': 3, 'four': 4, 'five': 5, 'six': 6, 'seven': 7,
            'eight': 8, 'nine': 9, 'ten': 10, 'eleven': 11, 'twelve': 12, 'thirteen': 13,
            'fourteen': 14, 'fifteen': 15, 'sixteen': 16, 'seventeen': 17, 'eighteen': 18,
            'nineteen': 19, 'twenty': 20, 'fifteeen': 15  # Handle typo in data
        }
        
        for row in deck_rows:
            did, week, unit, dstr, reading_list, topic1, topic2, monday_details, wednesday_details = row
            
            # Extract week number from week name (e.g., "Week Four - Monday" -> 4)
            week_num = None
            if week:
                # Try to find "Week <number>" or "Week <word>"
                match = re.search(r'Week\s+(\w+)', week, re.IGNORECASE)
                if match:
                    week_word = match.group(1).lower()
                    # Try as number first
                    if week_word.isdigit():
                        week_num = int(week_word)
                    # Try as word
                    elif week_word in week_words:
                        week_num = week_words[week_word]
            
            if not week_num:
                print(f"[WARN] Could not extract week number from '{week}', skipping deck {did}")
                continue
            
            if week_num not in deck_topics_by_week:
                deck_topics_by_week[week_num] = {
                    'monday_details': '',
                    'wednesday_details': '',
                    'dates': [],
                    'deck_ids': []
                }
            
            deck_topics_by_week[week_num]['deck_ids'].append(did)
            
            # Parse and add date
            if dstr:
                pd = parse_any_date(dstr)
                if pd:
                    deck_topics_by_week[week_num]['dates'].append(pd)
            
            # Build topics from topic1, topic2
            if 'Monday' in (week or '') or 'Mon' in (week or ''):
                topics = []
                if topic1 and topic1 not in ['Topic', '']:
                    topics.append(topic1.strip())
                if topic2 and topic2 not in ['Topic', '']:
                    topics.append(topic2.strip())
                
                # Only add monday_details if it's not a placeholder
                if monday_details and monday_details not in ['{NO CLASS}', 'Topic', '']:
                    topics.append(monday_details.strip())
                
                if topics:
                    deck_topics_by_week[week_num]['monday_details'] = '\n'.join(topics)
                    
            if 'Wednesday' in (week or '') or 'Wed' in (week or ''):
                topics = []
                if topic1 and topic1 not in ['Topic', '']:
                    topics.append(topic1.strip())
                if topic2 and topic2 not in ['Topic', '']:
                    topics.append(topic2.strip())
                
                # Only add wednesday_details if it's not a placeholder
                if wednesday_details and wednesday_details not in ['{NO CLASS}', 'Topic', '']:
                    topics.append(wednesday_details.strip())
                
                if topics:
                    deck_topics_by_week[week_num]['wednesday_details'] = '\n'.join(topics)
        
        # Build weeks from decks, will overlay CSV readings/holidays below
        print(f"[DEBUG] Creating {len(deck_topics_by_week)} weeks from decks...")
        
        # First, get all decks to assign sequential lecture numbers
        # We need to sort by actual date, not string representation
        c_decks = conn.cursor()
        c_decks.execute('''
            SELECT id, week, date, order_index 
            FROM decks 
            WHERE presentation_id = ? 
        ''', (presentation_id,))
        all_decks = c_decks.fetchall()
        
        # Parse dates and sort chronologically
        decks_with_dates = []
        for deck_id, week_name, date_str, order_idx in all_decks:
            parsed_date = parse_any_date(date_str)
            if parsed_date:
                decks_with_dates.append((deck_id, week_name, date_str, order_idx, parsed_date))
        
        # Sort by parsed date
        ordered_decks = sorted(decks_with_dates, key=lambda x: x[4])  # Sort by parsed_date
        
        # Build mapping of deck_id to lecture number and day
        lecture_words = ['ONE', 'TWO', 'THREE', 'FOUR', 'FIVE', 'SIX', 'SEVEN', 'EIGHT', 
                         'NINE', 'TEN', 'ELEVEN', 'TWELVE', 'THIRTEEN', 'FOURTEEN', 'FIFTEEN',
                         'SIXTEEN', 'SEVENTEEN', 'EIGHTEEN', 'NINETEEN', 'TWENTY',
                         'TWENTY-ONE', 'TWENTY-TWO', 'TWENTY-THREE', 'TWENTY-FOUR', 'TWENTY-FIVE',
                         'TWENTY-SIX', 'TWENTY-SEVEN', 'TWENTY-EIGHT', 'TWENTY-NINE', 'THIRTY']
        
        deck_lecture_info = {}  # deck_id -> (lecture_word, is_monday, is_wednesday)
        lecture_num = 0
        
        for deck_id, week_name, date_str, order_idx, parsed_date in ordered_decks:
            # Determine if this is a Monday or Wednesday deck
            is_monday = 'Monday' in (week_name or '') or 'Mon' in (week_name or '')
            is_wednesday = 'Wednesday' in (week_name or '') or 'Wed' in (week_name or '')
            
            # Check if this deck has actual topics in deck_topics_by_week
            has_content = False
            for week_num, week_data in deck_topics_by_week.items():
                if deck_id in week_data.get('deck_ids', []):
                    if is_monday and week_data.get('monday_details'):
                        has_content = True
                        break
                    if is_wednesday and week_data.get('wednesday_details'):
                        has_content = True
                        break
            
            if has_content:
                lecture_num += 1
                if lecture_num <= len(lecture_words):
                    deck_lecture_info[deck_id] = (lecture_words[lecture_num - 1], is_monday, is_wednesday)
                    print(f"[DEBUG] Deck {deck_id} ({week_name}, {date_str}, {parsed_date}): Lecture {lecture_words[lecture_num - 1]}")
        
        for week_num in sorted(deck_topics_by_week.keys()):
            deck_info = deck_topics_by_week[week_num]
            
            # Add lecture numbers based on deck order
            monday_details = deck_info['monday_details']
            wednesday_details = deck_info['wednesday_details']
            
            # Find which deck IDs contributed to this week and add their lecture numbers
            for deck_id in deck_info.get('deck_ids', []):
                if deck_id in deck_lecture_info:
                    lecture_word, is_monday, is_wednesday = deck_lecture_info[deck_id]
                    
                    # Add lecture number to the appropriate day
                    if is_monday and monday_details:
                        monday_details = lecture_word + '\n' + monday_details
                    if is_wednesday and wednesday_details:
                        wednesday_details = lecture_word + '\n' + wednesday_details
            
            # Assignments will be added from CSV below
            parsed_rows.append({
                'row': week_num,
                'unit': week_num,
                'dates': sorted(deck_info['dates']),
                'reading_list': '',  # Will be filled from CSV below
                'monday_details': monday_details,
                'wednesday_details': wednesday_details,
                'monday_no_class': False,  # Will be updated from CSV holidays below
                'wednesday_no_class': False,
                'assignment_tokens': [],
                'predicted_assignment_ids': [],
                'predicted_assignment_names': [],
                'predicted_assignment_dues': [],
                'predicted_assignment_shorts': [],
            })
        
        # CSV will be processed below to add readings and check for holidays
        rows_iter = []
    else:
        # Normal mode: not populate_from_db
        rows_iter = ws.iter_rows(values_only=True) if ws else []

    # Iterate rows; skip header rows heuristically if first row looks like headers
    assignments_to_create = []
    readings_by_week = {}
    assignments_by_week = {}  # Track assignments by week from CSV
    holidays_by_week = {}  # Track holidays with their week numbers
    if csv_rows:
        # Parse CSV as assignments/readings/holidays format
        # This runs even in populate_from_db mode to get readings and holidays
        for row in csv_rows[1:]:  # skip header
            if len(row) < 5:
                continue
            semester, week, assignment, due_date, xls_short = row[0], row[1], row[2], row[3], row[4]
            
            # Check for holidays/no class FIRST before treating as assignment
            if assignment in ['Holiday', 'No Class'] and due_date:
                # No class
                parsed = parse_any_date(due_date)
                if parsed:
                    no_class_map[parsed] = xls_short or assignment
                    # Also track by week number (deduplicate by date)
                    week_num = int(week) if week.isdigit() else None
                    if week_num:
                        # Check if this date is already in holidays_by_week for this week
                        existing = holidays_by_week.get(week_num, [])
                        if not any(d == parsed for d, _ in existing):
                            holidays_by_week.setdefault(week_num, []).append((parsed, xls_short or assignment))
            elif assignment and due_date:
                # Real assignment - track by week number
                week_num = int(week) if week.isdigit() else None
                if week_num:
                    assignments_by_week.setdefault(week_num, []).append({
                        'name': assignment,
                        'due_date': due_date,
                        'short': xls_short or assignment
                    })
                # Also create if not in populate_from_db mode
                if not populate_from_db:
                    assignments_to_create.append((assignment, due_date, xls_short or assignment))
            elif assignment and not due_date:
                # Reading for week
                week_num = int(week) if week.isdigit() else None
                if week_num:
                    readings_by_week.setdefault(week_num, []).append(assignment)
        
        # If populate_from_db, overlay CSV readings and holidays onto parsed_rows from decks
        if populate_from_db and (populate_template or populate_base):
            
            # Determine all weeks mentioned in CSV or that have decks
            all_weeks = set(deck_topics_by_week.keys())
            all_weeks.update(readings_by_week.keys())
            all_weeks.update(holidays_by_week.keys())
            
            # Create week entries for weeks that exist in CSV but not in decks
            existing_weeks = {pr['row'] for pr in parsed_rows}
            for week_num in all_weeks:
                if week_num not in existing_weeks:
                    # Create an empty week entry
                    parsed_rows.append({
                        'row': week_num,
                        'unit': week_num,
                        'dates': [],
                        'reading_list': '',
                        'monday_details': '',
                        'wednesday_details': '',
                        'monday_no_class': True,
                        'wednesday_no_class': True,
                        'assignment_tokens': [],
                        'predicted_assignment_ids': [],
                        'predicted_assignment_names': [],
                        'predicted_assignment_dues': [],
                        'predicted_assignment_shorts': [],
                    })
            
            # Sort by week number
            parsed_rows.sort(key=lambda x: x['row'])
            
            for pr in parsed_rows:
                week_num = pr['row']
                
                # Add readings
                if week_num in readings_by_week:
                    pr['reading_list'] = '\n'.join(readings_by_week[week_num])
                
                # Add assignments from CSV
                if week_num in assignments_by_week:
                    for assign in assignments_by_week[week_num]:
                        # Try to find matching assignment in database
                        c_assign = conn.cursor()
                        c_assign.execute('SELECT id, name, due_date, short FROM assignments WHERE name = ?', (assign['name'],))
                        match = c_assign.fetchone()
                        if match:
                            aid, aname, adue, ashort = match
                            pr['predicted_assignment_ids'].append(aid)
                            pr['predicted_assignment_names'].append(aname or '')
                            pr['predicted_assignment_dues'].append(adue or '')
                            pr['predicted_assignment_shorts'].append(ashort or assign['short'])
                        else:
                            # Assignment not in database, add with placeholder data
                            pr['assignment_tokens'].append(assign['name'])
                            pr['predicted_assignment_names'].append(assign['name'])
                            pr['predicted_assignment_dues'].append(assign['due_date'])
                            pr['predicted_assignment_shorts'].append(assign['short'])
                
                # Check dates for holidays and use holiday labels
                monday_holiday_labels = []
                wednesday_holiday_labels = []
                
                # First check if this week has holidays in holidays_by_week
                if week_num in holidays_by_week:
                    for date_str, label in holidays_by_week[week_num]:
                        dt = datetime.date.fromisoformat(date_str)
                        # Add date to dates list if not already there
                        if date_str not in pr['dates']:
                            pr['dates'].append(date_str)
                        if dt.weekday() == 0:  # Monday
                            pr['monday_no_class'] = True
                            monday_holiday_labels.append(label)
                        elif dt.weekday() == 2:  # Wednesday
                            pr['wednesday_no_class'] = True
                            wednesday_holiday_labels.append(label)
                
                # Also check deck dates for holidays (in case they weren't in CSV)
                for date_str in pr['dates']:
                    if date_str in no_class_map and date_str not in [h[0] for h in holidays_by_week.get(week_num, [])]:
                        dt = datetime.date.fromisoformat(date_str)
                        holiday_label = no_class_map[date_str]
                        if dt.weekday() == 0:  # Monday
                            pr['monday_no_class'] = True
                            monday_holiday_labels.append(holiday_label)
                        elif dt.weekday() == 2:  # Wednesday
                            pr['wednesday_no_class'] = True
                            wednesday_holiday_labels.append(holiday_label)
                
                # If no deck topics but have holiday labels, use the labels with empty lines around them
                if not pr['monday_details'] and monday_holiday_labels:
                    # Empty line, label, empty line (3 rows total, label in middle)
                    pr['monday_details'] = '' + '\n'.join([''] + monday_holiday_labels + [''])
                    pr['monday_no_class'] = True
                elif not pr['monday_details']:
                    pr['monday_no_class'] = True
                    
                if not pr['wednesday_details'] and wednesday_holiday_labels:
                    # Empty line, label, empty line (3 rows total, label in middle)
                    pr['wednesday_details'] = '' + '\n'.join([''] + wednesday_holiday_labels + [''])
                    pr['wednesday_no_class'] = True
                elif not pr['wednesday_details']:
                    pr['wednesday_no_class'] = True
                
                if week_num == 4:
                    print(f"[DEBUG] Week 4: wednesday_details={repr(pr['wednesday_details'])}, wednesday_no_class={pr['wednesday_no_class']}")
                    
                print(f"[DEBUG]   Week {week_num}: {len(pr['dates'])} dates, monday={bool(pr['monday_details'])}, wednesday={bool(pr['wednesday_details'])}, {len(pr['predicted_assignment_ids'])} assignments, readings={len(pr['reading_list'].split(chr(10)) if pr['reading_list'] else [])}")
        
        rows_iter = []  # no standard rows to process
    else:
        if populate_from_db and (populate_template or populate_base):
            rows_iter = []
        else:
            rows_iter = list(ws.iter_rows(values_only=True))
    for i, row in enumerate(rows_iter, start=1):
        # Skip empty rows
        if not any(cell not in (None, "") for cell in row[:5]):
            continue
        # Skip header row if it contains non-data header tokens
        first = row[0] or ''
        if i == 1 and isinstance(first, str) and any(k in first.lower() for k in ('unit', 'week', 'reading')):
            continue

        parsed = parse_xls_row(row, semester=semester)

        # Row-level predicted assignments (for export/template autopopulation and reporting)
        row_predicted_ids: List[int] = []
        row_predicted_names: List[str] = []
        row_predicted_dues: List[str] = []
        row_predicted_shorts: List[str] = []
        if semester and parsed['assignments']:
            for token in parsed['assignments']:
                candidates = find_assignment_candidates(conn, token, semester)
                for aid, name in candidates:
                    row_predicted_ids.append(aid)
                    row_predicted_names.append(name)
                    cursor.execute('SELECT due_date, short FROM assignments WHERE id = ?', (aid,))
                    rowr = cursor.fetchone()
                    if rowr:
                        row_predicted_dues.append(rowr[0] or '')
                        row_predicted_shorts.append(rowr[1] or '')

        # Add parsed row data for template export
        parsed_rows.append({
            'row': i,
            'first': first,
            'unit': parsed['unit'],
            'dates': parsed['dates'],
            'reading_list': parsed['reading_list'] or '',
            'monday_details': parsed['monday_details'] or '',
            'wednesday_details': parsed['wednesday_details'] or '',
            'assignment_tokens': parsed['assignments'],
            'predicted_assignment_ids': row_predicted_ids,
            'predicted_assignment_names': row_predicted_names,
            'predicted_assignment_dues': row_predicted_dues,
            'predicted_assignment_shorts': row_predicted_shorts,
        })

        if not parsed['dates']:
            # If no date could be parsed, skip (still included in parsed_rows for template)
            print(f"Skipping row {i}: no date found in '{first}'")
            skipped += 1
            skipped_rows.append({'row': i, 'first': first})
            continue

        for date_iso in parsed['dates']:
            deck_ids = find_deck_ids_by_date(conn, date_iso)
            if not deck_ids:
                if create_decks:
                    # Create a new deck row with minimal info
                    print(f"Creating deck for date {date_iso}")
                    cursor.execute('INSERT INTO decks (date, week) VALUES (?, ?)', (date_iso, f"Week of {date_iso}"))
                    conn.commit()
                    deck_id = cursor.lastrowid
                    deck_ids = [deck_id]
                else:
                    print(f"No deck found for date {date_iso}; skipping")
                    skipped += 1
                    skipped_rows.append({'row': i, 'first': first, 'date': date_iso})
                    continue

            # First, select deck dates for candidate decks (to support reading/week logic)
            deck_dates = _select_deck_dates(conn, deck_ids)

            for deck_id in deck_ids:
                # Update deck fields (unit, monday/wednesday). Handle readings separately based on apply_readings.
                reported_updates = {}
                if parsed['unit'] is not None:
                    reported_updates['unit'] = parsed['unit']
                if parsed['reading_list'] is not None:
                    # Include reading_list in report for context; but only schedule write operations per apply_readings
                    if apply_readings == 'none':
                        reported_updates['reading_list'] = parsed['reading_list']
                if parsed['monday_details'] is not None:
                    reported_updates['monday_details'] = parsed['monday_details']
                if parsed['wednesday_details'] is not None:
                    reported_updates['wednesday_details'] = parsed['wednesday_details']

                # Predict assignment matches (read-only) for report
                predicted_ids = []
                predicted_names = []
                predicted_dues = []
                if link_assignments and parsed['assignments'] and semester:
                    for token in parsed['assignments']:
                        candidates = find_assignment_candidates(conn, token, semester)
                        for aid, name in candidates:
                            predicted_ids.append(aid)
                            predicted_names.append(name)
                            # fetch due_date for better reporting
                            c.execute('SELECT due_date FROM assignments WHERE id = ?', (aid,))
                            row_due = c.fetchone()
                            predicted_dues.append(row_due[0] if row_due and row_due[0] else '')

                planned_actions.append({
                    'row': i,
                    'first': first,
                    'date_iso': date_iso,
                    'deck_id': deck_id,
                    'updates': reported_updates,
                    'assignment_tokens': parsed['assignments'],
                    'predicted_assignment_ids': predicted_ids,
                    'predicted_assignment_names': predicted_names,
                    'predicted_assignment_dues': predicted_dues,
                })

                # Apply non-reading updates immediately (unit, monday_details, wednesday_details)
                apply_updates = {k: v for k, v in reported_updates.items() if k != 'reading_list'}
                if apply_updates:
                        set_clause = ', '.join([f"{k} = ?" for k in apply_updates.keys()])
                        params = list(apply_updates.values()) + [deck_id]
                        print(f"Update deck {deck_id} for date {date_iso}: set {apply_updates}")
                        if not dry_run:
                            cursor.execute(f'UPDATE decks SET {set_clause} WHERE id = ?', params)
                            conn.commit()
                        updated_deck_ids.add(deck_id)
                # Collect reading_list items for later aggregate application depending on apply_readings
                if parsed['reading_list'] and apply_readings != 'none':
                    items = split_readings(parsed['reading_list'])
                    if apply_readings == 'all':
                        # apply to all candidate decks (accumulate per deck)
                        for did in deck_ids:
                            lst = deck_readings.setdefault(did, [])
                            for it in items:
                                if it not in lst:
                                    lst.append(it)
                    elif apply_readings in ('first', 'monday'):
                        # choose a single deck among deck_ids for this row
                        # prefer monday if requested
                        chosen = None
                        if apply_readings == 'monday':
                            for did in deck_ids:
                                d_iso = deck_dates.get(did)
                                if d_iso:
                                    dow = datetime.date.fromisoformat(d_iso).weekday()
                                    if dow == 0:
                                        chosen = did
                                        break
                        if chosen is None:
                            # fallback to earliest date among deck_ids
                            earliest = None
                            for did in deck_ids:
                                d_iso = deck_dates.get(did)
                                if d_iso:
                                    d_obj = datetime.date.fromisoformat(d_iso)
                                    if earliest is None or d_obj < earliest[0]:
                                        earliest = (d_obj, did)
                            if earliest:
                                chosen = earliest[1]
                        if chosen:
                            lst = deck_readings.setdefault(chosen, [])
                            for it in items:
                                if it not in lst:
                                    lst.append(it)
                    elif apply_readings == 'per_week':
                        # aggregate by ISO week key for each deck/date
                        for did in deck_ids:
                            d_iso = deck_dates.get(did)
                            if not d_iso:
                                continue
                            d_obj = datetime.date.fromisoformat(d_iso)
                            week_key = (d_obj.isocalendar()[0], d_obj.isocalendar()[1])
                            entry = weekly_readings.setdefault(week_key, {'deck_id': None, 'date_iso': d_iso, 'readings': []})
                            # set canonical deck if not yet set (prefer Monday, else earliest)
                            if entry['deck_id'] is None:
                                # try find monday among deck_ids for this week
                                chosen = None
                                for dd in deck_ids:
                                    d2_iso = deck_dates.get(dd)
                                    if not d2_iso:
                                        continue
                                    if datetime.date.fromisoformat(d2_iso).weekday() == 0:
                                        chosen = dd
                                        break
                                if chosen is None:
                                    # pick earliest
                                    earliest = None
                                    for dd in deck_ids:
                                        d2_iso = deck_dates.get(dd)
                                        if not d2_iso:
                                            continue
                                        d2 = datetime.date.fromisoformat(d2_iso)
                                        if earliest is None or d2 < earliest[0]:
                                            earliest = (d2, dd)
                                    if earliest:
                                        chosen = earliest[1]
                                entry['deck_id'] = chosen or did
                                entry['date_iso'] = d_iso
                            # append items preserving order
                            for it in items:
                                if it not in entry['readings']:
                                    entry['readings'].append(it)

                # Link assignments if requested
                if link_assignments and parsed['assignments']:
                    if not semester:
                        raise RuntimeError('When linking assignments you must provide --semester')
                    for aid_token in parsed['assignments']:
                        # On dry_run, we already predicted IDs; on apply, perform actual link
                        if not dry_run:
                            linked = link_assignment_to_deck(conn, deck_id, aid_token, semester)
                            if linked:
                                for aid in linked:
                                    print(f"Linked assignment {aid} to deck {deck_id}")
                                    linked_assignments += 1
                            else:
                                print(f"Assignment '{aid_token}' not found for semester {semester}; skipping link")
                        else:
                            # dry-run: count predicted ids
                            candidates = find_assignment_candidates(conn, aid_token, semester)
                            if candidates:
                                linked_assignments += len(candidates)


    # Apply aggregated reading_list updates (per deck or per week) after scanning all rows
    def _apply_deck_readings_map(dr_map: Dict[int, List[str]]):
        # update counts using `updated_deck_ids` set
        for did, readings in dr_map.items():
            # Read current reading_list
            c.execute('SELECT reading_list, date FROM decks WHERE id = ?', (did,))
            row = c.fetchone()
            if not row:
                continue
            cur_text = row[0] if row and row[0] else ''
            cur_lines = split_readings(cur_text) if cur_text else []
            to_append = [r for r in readings if r not in cur_lines]
            if not to_append:
                continue
            new_text = cur_text + ('\n' + '\n'.join(to_append) if cur_text else '\n'.join(to_append))
            # Add planned action
            planned_actions.append({
                'row': f'reading_deck_{did}',
                'first': '',
                'date_iso': row[1] if row and row[1] else '',
                'deck_id': did,
                'updates': {'reading_list': new_text},
                'assignment_tokens': [],
                'predicted_assignment_ids': [],
                'predicted_assignment_names': [],
            })
            print(f"Add reading to deck {did}: '{'; '.join(to_append)}'")
            if not dry_run:
                c.execute('UPDATE decks SET reading_list = ? WHERE id = ?', (new_text, did))
                conn.commit()
            updated_deck_ids.add(did)

    # Merge weekly_readings into a deck_readings map (choose deck per week)
    for (y, w), entry in weekly_readings.items():
        did = entry.get('deck_id')
        if not did:
            continue
        lst = deck_readings.setdefault(did, [])
        for r in entry.get('readings', []):
            if r not in lst:
                lst.append(r)

    # Apply deck_readings (aggregated from per-row or weekly)
    _apply_deck_readings_map(deck_readings)
    # Compute unique deck update count
    updated_decks = len(updated_deck_ids)

    # If requested, optionally link assignments by due date (opt-in)
    if link_by_due and semester and not dry_run:
        added = _link_assignments_by_due(semester, link_by_due_n)
        linked_assignments += added

    # If requested, write report CSV (deck-level proposals)
    if report_csv:
        with open(report_csv, 'w', newline='') as rf:
            fieldnames = ['row', 'first', 'date_iso', 'deck_id', 'updates_json', 'assignment_tokens', 'predicted_assignment_ids', 'predicted_assignment_names', 'predicted_assignment_dues']
            w = csv.DictWriter(rf, fieldnames=fieldnames)
            w.writeheader()
            for act in planned_actions:
                w.writerow({
                    'row': act['row'],
                    'first': act['first'],
                    'date_iso': act['date_iso'],
                    'deck_id': act['deck_id'],
                    'updates_json': json.dumps(act['updates'], ensure_ascii=False),
                    'assignment_tokens': '|'.join(act['assignment_tokens']) if act['assignment_tokens'] else '',
                    'predicted_assignment_ids': '|'.join(str(x) for x in act.get('predicted_assignment_ids', [])),
                    'predicted_assignment_names': '|'.join(act.get('predicted_assignment_names', [])) if act.get('predicted_assignment_names') else '',
                    'predicted_assignment_dues': '|'.join(act.get('predicted_assignment_dues', [])) if act.get('predicted_assignment_dues') else '',
                })
        print(f"Report written: {report_csv}")

    # If requested, write a CSV to populate the XLS template (row-by-row, no DB writes)
    if export_template_csv:
        with open(export_template_csv, 'w', newline='') as ef:
            fieldnames = ['row', 'first', 'dates', 'reading_list', 'monday_details', 'wednesday_details', 'assignment_tokens', 'predicted_shorts', 'predicted_names', 'predicted_ids', 'predicted_dues']
            w2 = csv.DictWriter(ef, fieldnames=fieldnames)
            w2.writeheader()
            for r in parsed_rows:
                w2.writerow({
                    'row': r['row'],
                    'first': r['first'],
                    'dates': '|'.join(r['dates']) if r['dates'] else '',
                    'reading_list': r['reading_list'],
                    'monday_details': r['monday_details'],
                    'wednesday_details': r['wednesday_details'],
                    'assignment_tokens': '|'.join(r['assignment_tokens']) if r['assignment_tokens'] else '',
                    'predicted_shorts': '|'.join(r.get('predicted_assignment_shorts', [])) if r.get('predicted_assignment_shorts') else '',
                    'predicted_names': '|'.join(r.get('predicted_assignment_names', [])) if r.get('predicted_assignment_names') else '',
                    'predicted_ids': '|'.join(str(x) for x in r.get('predicted_assignment_ids', [])) if r.get('predicted_assignment_ids') else '',
                    'predicted_dues': '|'.join(r.get('predicted_assignment_dues', [])) if r.get('predicted_assignment_dues') else '',
                })
        print(f"Template export written: {export_template_csv}")

    # Helper: find a main sheet that contains placeholders like {DATES} or {readings}
    def _find_main_sheet(wb):
        for name in wb.sheetnames:
            ws = wb[name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if isinstance(cell, str):
                        up = cell.upper()
                        if '{DATES}' in up or '{READ' in up or '{UNIT' in up or '{WEEK' in up:
                            return name
        return None

    # Helper: copy formatting (styles + merges) from a source row to a target row
    def _copy_row_format(ws, src_row, tgt_row, max_col=None):
        try:
            from openpyxl.utils import get_column_letter
        except Exception:
            def get_column_letter(i):
                return str(i)
        if max_col is None:
            max_col = ws.max_column
        for c in range(1, max_col + 1):
            src = ws.cell(row=src_row, column=c)
            tgt = ws.cell(row=tgt_row, column=c)
            try:
                tgt.font = copy(src.font)
                tgt.border = copy(src.border)
                tgt.fill = copy(src.fill)
                tgt.number_format = src.number_format
                tgt.protection = copy(src.protection)
                tgt.alignment = copy(src.alignment)
            except Exception:
                # best-effort - ignore copying issues
                pass
        # replicate merged cell ranges that include the source row
        merges = list(ws.merged_cells.ranges)
        for m in merges:
            try:
                if m.min_row <= src_row <= m.max_row:
                    row_offset = tgt_row - src_row
                    new_min_row = m.min_row + row_offset
                    new_max_row = m.max_row + row_offset
                    new_range = f"{get_column_letter(m.min_col)}{new_min_row}:{get_column_letter(m.max_col)}{new_max_row}"
                    # Only add merge if it doesn't exist
                    if new_range not in [str(r) for r in ws.merged_cells.ranges]:
                        ws.merge_cells(new_range)
            except Exception:
                continue

    # Process CSV assignments and readings
    if assignments_to_create:
        print(f"[DEBUG] Creating {len(assignments_to_create)} assignments")
        if not dry_run:
            for name, due, short in assignments_to_create:
                cursor.execute('INSERT INTO assignments (semester, name, due_date, short) VALUES (?, ?, ?, ?)', (semester, name, due, short))
            conn.commit()
        if link_assignments:
            _link_assignments_by_due(semester, link_by_due_n)
    if readings_by_week:
        print(f"[DEBUG] Applying readings for weeks {list(readings_by_week.keys())}")
        if not dry_run:
            for week, readings in readings_by_week.items():
                reading_list = '|'.join(readings)
                cursor.execute('UPDATE decks SET reading_list = ? WHERE unit = ?', (reading_list, str(week)))
            conn.commit()

    # If requested, populate a template workbook (e.g., .xltx) or use an existing base workbook and save as a clean .xlsx
    if populate_template or populate_base:
        base_path = populate_base or populate_template
        out_path = populate_output or os.path.join('/tmp', f"{os.path.splitext(os.path.basename(base_path))[0]}_populated.xlsx")
        print(f"Populating template {base_path} -> {out_path}")
        try:
            t_wb = load_workbook(base_path)
        except Exception as e:
            print(f"Cannot load template/base workbook {base_path}: {e}")
        else:
            # Helper function to safely set cell values (handles merged cells)
            def safe_set_value(ws, row, col, value):
                from openpyxl.cell.cell import MergedCell
                cell = ws.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    # Find the top-left cell of the merged range
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            top_left_cell.value = value
                            return
                else:
                    cell.value = value
            target_sheet = None
            if semester:
                for cand in (f'{semester} Dates', semester):
                    if cand in t_wb.sheetnames:
                        target_sheet = cand
                        break
            if not target_sheet:
                for name in t_wb.sheetnames:
                    if 'dates' in name.lower():
                        target_sheet = name
                        break
            if not target_sheet:
                print(f"No 'Dates' sheet found in template/base {base_path}; using active sheet")
                target_sheet = t_wb.active.title
            
            ws_dates = t_wb[target_sheet]

            # Import 'Holiday'/'No Class' entries from the data sheet (if present) into no_class_map
            sheet_holiday_count = 0
            try:
                for row in ws_dates.iter_rows(values_only=True):
                    if not any(row):
                        continue
                    for idx, cell in enumerate(row):
                        if not cell or not isinstance(cell, str):
                            continue
                        token = cell.strip().lower()
                        if token in ('holiday', 'no class'):
                                # find date token (prefer the next column)
                                date_candidate = None
                                if idx + 1 < len(row) and row[idx + 1]:
                                    date_candidate = row[idx + 1]
                                # fallback: search any cell in the row for a parsable date
                                if not parse_any_date(date_candidate):
                                    for cell_cand in row:
                                        if cell_cand and parse_any_date(cell_cand):
                                            date_candidate = cell_cand
                                            break
                                date_iso = parse_any_date(date_candidate)
                                if not date_iso:
                                    continue
                                # find a short/label (prefer next next column, then last textual column)
                                short_label = None
                                if idx + 2 < len(row) and row[idx + 2]:
                                    short_label = str(row[idx + 2]).strip()
                                if not short_label:
                                    # find last non-marker textual cell that's not the token itself
                                    for j in range(len(row) - 1, -1, -1):
                                        cval = row[j]
                                        if cval and isinstance(cval, str) and cval.strip().lower() not in ('holiday', 'no class'):
                                            short_label = cval.strip()
                                            break
                                short_label = short_label or token.title()
                                no_class_map[date_iso] = short_label
                                sheet_holiday_count += 1
                if sheet_holiday_count:
                    print(f"Loaded {sheet_holiday_count} no-class entries from '{target_sheet}' sheet")
            except Exception:
                # non-fatal; continue without sheet-derived holidays
                pass

            rows_to_write = []
            if semester:
                # Prefer ordering from parsed_rows (calendar order) so the 'Dates' sheet preserves the workbook's sequence
                seen = set()
                for r in parsed_rows:
                    names = r.get('predicted_assignment_names', [])
                    shorts = r.get('predicted_assignment_shorts', [])
                    dues = r.get('predicted_assignment_dues', [])
                    for i, nm in enumerate(names):
                        short = shorts[i] if i < len(shorts) else ''
                        due = dues[i] if i < len(dues) else ''
                        key = (nm, due, short)
                        if key not in seen:
                            seen.add(key)
                            rows_to_write.append(key)
                # If the parsed_rows produced no entries (e.g., template file had no assignment tokens),
                # fall back to DB listing ordered by due_date, name for a reasonable default ordering.
                if not rows_to_write:
                    c.execute('SELECT name, due_date, short FROM assignments WHERE semester = ? ORDER BY due_date, name', (semester,))
                    for name, due, short in c.fetchall():
                        rows_to_write.append((name, due, short))
            else:
                seen = set()
                for r in parsed_rows:
                    names = r.get('predicted_assignment_names', [])
                    shorts = r.get('predicted_assignment_shorts', [])
                    dues = r.get('predicted_assignment_dues', [])
                    for i, nm in enumerate(names):
                        short = shorts[i] if i < len(shorts) else ''
                        due = dues[i] if i < len(dues) else ''
                        key = (nm, due, short)
                        if key not in seen:
                            seen.add(key)
                            rows_to_write.append(key)
            # Build placeholder row info: map each placeholder row to the columns that hold {DATES}, {WEEK}, {readings}, {ASSIGNMENT}
            main_sheet_name = _find_main_sheet(t_wb)
            if main_sheet_name:
                main_ws = t_wb[main_sheet_name]
            
            print(f"[DEBUG] Identifying template block types...")
            # The template has 4 block types that we need to identify and use:
            # Block 1: Normal (both Monday & Wednesday) - typically rows with UNIT #
            # Block 2: No Monday class - monday_topics missing from row 1 of block
            # Block 3: No Wednesday class - wednesday_topics missing from row 1 of block
            # Block 4: No class both days - both topics missing from row 1 of block
            
            # Find blocks by scanning for {WEEK #} placeholders
            template_blocks = []
            for rr_idx, row in enumerate(main_ws.iter_rows(values_only=True), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    if isinstance(cell, str) and '{WEEK #}' in cell:
                        # Found start of a block, analyze it
                        block_rows = [rr_idx, rr_idx + 1, rr_idx + 2]  # 3-row blocks
                        
                        # Check if row before has UNIT #
                        has_unit = False
                        if rr_idx > 1:
                            prev_row = list(main_ws.iter_rows(min_row=rr_idx-1, max_row=rr_idx-1, values_only=True))[0]
                            has_unit = any(isinstance(c, str) and '{UNIT #}' in c for c in prev_row)
                            if has_unit:
                                block_rows.insert(0, rr_idx - 1)
                        
                        # Determine block type by checking first row with placeholders
                        first_placeholder_row = list(main_ws.iter_rows(min_row=rr_idx, max_row=rr_idx, values_only=True))[0]
                        has_monday = any(isinstance(c, str) and '{monday_topics}' in c for c in first_placeholder_row)
                        has_wednesday = any(isinstance(c, str) and '{wednesday_topics}' in c for c in first_placeholder_row)
                        
                        block_type = 'normal' if has_monday and has_wednesday else \
                                   'no_monday' if not has_monday and has_wednesday else \
                                   'no_wednesday' if has_monday and not has_wednesday else \
                                   'no_class'
                        
                        template_blocks.append({
                            'type': block_type,
                            'rows': block_rows,
                            'has_unit': has_unit
                        })
                        print(f"[DEBUG]   Found {block_type} block at rows {block_rows}")
                        break
            
            if not template_blocks:
                print("[ERROR] No template blocks found with {WEEK #} placeholders")
                return
            
            # Create a new workbook by copying header and populating from template blocks
            from openpyxl.utils import get_column_letter
            from copy import copy
            
            # Use the main worksheet and populate in place
            output_ws = main_ws
            header_row = 1
            
            # Store template block data before we start modifying
            template_data = {}
            for tb in template_blocks:
                template_data[tb['type']] = []
                for row_idx in tb['rows']:
                    row_values = []
                    for col in range(1, 10):
                        cell = main_ws.cell(row=row_idx, column=col)
                        row_values.append({
                            'value': cell.value,
                            'font': copy(cell.font) if cell.has_style else None,
                            'border': copy(cell.border) if cell.has_style else None,
                            'fill': copy(cell.fill) if cell.has_style else None,
                            'number_format': copy(cell.number_format) if cell.has_style else None,
                            'alignment': copy(cell.alignment) if cell.has_style else None
                        })
                    template_data[tb['type']].append(row_values)
            
            # Now clear the template blocks (but keep header)
            max_template_row = max(max(tb['rows']) for tb in template_blocks)
            output_ws.delete_rows(header_row + 1, max_template_row)
            
            current_output_row = header_row + 1
            parsed_for_main = [pr for pr in parsed_rows if pr.get('dates')]
            
            print(f"[DEBUG] Populating {len(parsed_for_main)} weeks...")
            
            def copy_template_row(ws, dest_row, template_row_data):
                """Copy a template row including values and formatting"""
                for col_idx, cell_data in enumerate(template_row_data, start=1):
                    dest_cell = ws.cell(row=dest_row, column=col_idx)
                    dest_cell.value = cell_data['value']
                    if cell_data['font']:
                        dest_cell.font = copy(cell_data['font'])
                    if cell_data['border']:
                        dest_cell.border = copy(cell_data['border'])
                    if cell_data['fill']:
                        dest_cell.fill = copy(cell_data['fill'])
                    if cell_data['number_format']:
                        dest_cell.number_format = cell_data['number_format']
                    if cell_data['alignment']:
                        dest_cell.alignment = copy(cell_data['alignment'])
            
            def num_to_words(n):
                words = ['', 'One', 'Two', 'Three', 'Four', 'Five', 'Six', 'Seven', 'Eight', 'Nine', 'Ten',
                        'Eleven', 'Twelve', 'Thirteen', 'Fourteen', 'Fifteen', 'Sixteen']
                return words[n] if 0 < n < len(words) else str(n)
            
            for week_idx, pr in enumerate(parsed_for_main, start=1):
                # Determine which template block to use
                has_monday = bool(pr.get('monday_details'))
                has_wednesday = bool(pr.get('wednesday_details'))
                
                block_type = 'normal' if has_monday and has_wednesday else \
                           'no_monday' if not has_monday and has_wednesday else \
                           'no_wednesday' if has_monday and not has_wednesday else \
                           'no_class'
                
                # Get template block data
                if block_type not in template_data:
                    print(f"[WARN] No template data for type {block_type}, using first available block type")
                    # Use the first available block type as fallback
                    if template_data:
                        block_type = list(template_data.keys())[0]
                    else:
                        print("[ERROR] No template blocks available")
                        continue
                
                block_template_rows = template_data[block_type]
                print(f"[DEBUG]   Week {week_idx}: {block_type} block, {len(block_template_rows)} rows")
                
                # Copy template block rows to output and populate
                block_start_row = current_output_row
                
                # Format date helper
                def _fmt_date(tok):
                    try:
                        d_obj = datetime.date.fromisoformat(tok)
                        return f"{d_obj.month}/{d_obj.day}"
                    except Exception:
                        return str(tok)
                
                # Populate each row in the block
                # Track which line of topics we're on (separate from row offset)
                monday_topic_line = 0
                wednesday_topic_line = 0
                reading_line = 0
                assignment_line = 0
                
                for offset, template_row_data in enumerate(block_template_rows):
                    dest_row = block_start_row + offset
                    
                    # First copy the template row
                    try:
                        copy_template_row(output_ws, dest_row, template_row_data)
                    except Exception as e:
                        print(f"[ERROR] Week {week_idx}, offset {offset}: copy_template_row failed: {e}")
                        raise
                    
                    # Now replace placeholders in the copied row
                    try:
                        for col_idx, cell_data in enumerate(template_row_data, start=1):
                            cell_val = cell_data['value']
                            if not isinstance(cell_val, str):
                                continue
                            
                            # Replace placeholders with actual data
                            new_val = None
                            apply_red_bg = False
                            
                            if '{UNIT #}' in cell_val:
                                new_val = week_idx
                            elif '{WEEK #}' in cell_val:
                                new_val = num_to_words(week_idx)
                            elif '{DATES}' in cell_val:
                                all_dates = pr.get('dates', [])
                                if len(all_dates) == 1:
                                    new_val = _fmt_date(all_dates[0])
                                elif len(all_dates) > 1:
                                    new_val = f"{_fmt_date(all_dates[0])}-{_fmt_date(all_dates[-1])}"
                            elif '{readings}' in cell_val:
                                reading_lines = pr.get('reading_list', '').split('\n') if pr.get('reading_list') else []
                                if reading_lines and reading_line < len(reading_lines):
                                    new_val = reading_lines[reading_line]
                                    reading_line += 1
                            elif '{monday_topics}' in cell_val:
                                # Check if there are actual details (like holiday labels)
                                monday_details = pr.get('monday_details', '')
                                monday_lines = monday_details.split('\n') if monday_details else []
                                
                                if monday_lines and monday_topic_line < len(monday_lines):
                                    # Use the actual detail (topic or holiday label)
                                    new_val = monday_lines[monday_topic_line]
                                    monday_topic_line += 1
                                
                                # Apply red background if this is marked as no-class (all rows get red)
                                if pr.get('monday_no_class'):
                                    apply_red_bg = True
                                        
                            elif '{wednesday_topics}' in cell_val:
                                # Check if there are actual details (like holiday labels)
                                wednesday_details = pr.get('wednesday_details', '')
                                wednesday_lines = wednesday_details.split('\n') if wednesday_details else []
                                
                                if wednesday_lines and wednesday_topic_line < len(wednesday_lines):
                                    # Use the actual detail (topic or holiday label)
                                    new_val = wednesday_lines[wednesday_topic_line]
                                    wednesday_topic_line += 1
                                
                                # Apply red background if this is marked as no-class (all rows get red)
                                if pr.get('wednesday_no_class'):
                                    apply_red_bg = True
                                        
                            elif '{ASSIGNMENT}' in cell_val:
                                pshorts = pr.get('predicted_assignment_shorts', [])
                                if pshorts and assignment_line < len(pshorts):
                                    new_val = pshorts[assignment_line]
                                    assignment_line += 1
                            
                            # Only update if we found a replacement
                            if new_val is not None:
                                safe_set_value(output_ws, dest_row, col_idx, new_val)
                                
                                # Apply red background if this is a no-class day
                                if apply_red_bg:
                                    from openpyxl.styles import PatternFill
                                    cell = output_ws.cell(row=dest_row, column=col_idx)
                                    cell.fill = PatternFill(start_color='BE5014', end_color='BE5014', fill_type='solid')
                    except Exception as e:
                        print(f"[ERROR] Week {week_idx}, offset {offset}, col {col_idx}: {e}")
                        print(f"  monday_details: {repr(pr.get('monday_details', ''))}")
                        print(f"  wednesday_details: {repr(pr.get('wednesday_details', ''))}")
                        import traceback
                        traceback.print_exc()
                        raise
                    
                    current_output_row += 1
                
                # Add blank row after block
                current_output_row += 1
            
            try:
                # If this was a DB-driven populate, clear any leftover placeholder tokens like '{...}'
                # If this was a DB-driven populate, clear any leftover placeholder tokens like '{...}'
                # so the output sheet doesn't contain raw template markers.
                if populate_from_db:
                    import re
                    for wsname in t_wb.sheetnames:
                        ws_tmp = t_wb[wsname]
                        for r in ws_tmp.iter_rows():
                            for cell in r:
                                v = cell.value
                                if isinstance(v, str) and '{' in v and '}' in v:
                                    cell.value = None
                t_wb.save(out_path)
            except Exception as e:
                print(f"Error saving populated workbook: {e}")
            else:
                print(f"Saved populated workbook to {out_path}")
            try:
                import zipfile, shutil, tempfile
                with zipfile.ZipFile(out_path, 'r') as zin:
                    ct = zin.read('[Content_Types].xml').decode('utf-8')
                    if 'template.main+xml' in ct:
                        tmp_path = out_path + '.fixed'
                        with zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
                            for info in zin.infolist():
                                data = zin.read(info.filename)
                                if info.filename == '[Content_Types].xml':
                                    s = data.decode('utf-8').replace('application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml')
                                    data = s.encode('utf-8')
                                zout.writestr(info, data)
                        shutil.move(tmp_path, out_path)
                        print(f"Fixed content type for {out_path}")
            except Exception as e:
                print(f"Error fixing content type for {out_path}: {e}")

    conn.close()
    if dry_run:
        print(f"Dry run: {updated_decks} deck updates, {linked_assignments} links, {skipped} skipped rows")
    else:
        print(f"Applied: {updated_decks} deck updates, {linked_assignments} links, {skipped} skipped rows")

    return updated_decks, linked_assignments, skipped


def split_readings(s: str) -> List[str]:
    """Split a reading_list cell into multiple reading items, preserving order and trimming whitespace."""
    if not s:
        return []
    parts = re.split(r"[\n;|]+", s)
    items: List[str] = []
    for p in parts:
        t = p.strip()
        if t and t not in items:
            items.append(t)
    return items


def _select_deck_dates(conn: sqlite3.Connection, ids: List[int]) -> Dict[int, Optional[str]]:
    """Return mapping of deck_id -> normalized ISO date or None"""
    if not ids:
        return {}
    c = conn.cursor()
    placeholders = ','.join('?' for _ in ids)
    c.execute(f'SELECT id, date FROM decks WHERE id IN ({placeholders})', ids)
    res: Dict[int, Optional[str]] = {}
    for did, d in c.fetchall():
        res[did] = parse_any_date(d)
    return res



def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description='Import calendar rows from an Excel template to decks')
    parser.add_argument('--xls', required=True, help='Path to Excel (.xlsx) file')
    parser.add_argument('--db', default='presentations.db', help='Path to SQLite DB')
    parser.add_argument('--semester', help='Semester code to resolve assignment short codes (e.g., SP2026)')
    parser.add_argument('--backup', action='store_true', help='Create a timestamped backup of the DB')
    parser.add_argument('--dry-run', action='store_true', help='Do not modify DB, just show actions')
    parser.add_argument('--link-assignments', action='store_true', help='Link assignments using short codes in column E (requires --semester)')
    parser.add_argument('--link-by-due', action='store_true', help='Link assignments to the prior N decks by assignment due date (requires --semester)')
    parser.add_argument('--link-by-due-n', type=int, default=3, help='Number of prior decks to link per assignment when using --link-by-due (default 3)')
    parser.add_argument('--create-decks', action='store_true', help='Create deck rows for dates that are missing')
    parser.add_argument('--report-csv', help='Write a CSV report of proposed updates and predicted assignment links (use with --dry-run)')
    parser.add_argument('--export-template-csv', help='Write a CSV suitable for populating the XLS template with assignment shorts and reading lists (no DB writes)')
    parser.add_argument('--populate-template', help='Path to the template workbook (e.g., .xltx) to populate and save as .xlsx')
    parser.add_argument('--populate-base', help='Path to an existing master workbook (.xlsx) to use as the base for populated output')
    parser.add_argument('--populate-output', help='Path to write the populated .xlsx (defaults to /tmp/<basename>_populated.xlsx)')
    parser.add_argument('--no-class-csv', help='CSV of date,label for days with no class (e.g., MLK Day)')
    parser.add_argument('--apply-readings', choices=['none','first','per_week','all','monday'], default='none', help='How to apply reading_list cells to decks')
    parser.add_argument('--populate-from-db', action='store_true', help='When populating a template, use decks from the DB only (ignore rows in input xls)')
    parser.add_argument('--presentation-id', type=int, help='When used with --populate-from-db, restrict to decks for this presentation id')
    parser.add_argument('--clear-semester-data', action='store_true', help='Clear all assignments, links, and readings for the semester before importing (makes XLS authoritative)')
    args = parser.parse_args(argv)

    try:
        import_calendar_xls(
            args.xls,
            db_path=args.db,
            semester=args.semester,
            backup=args.backup,
            dry_run=args.dry_run,
            link_assignments=args.link_assignments,
            link_by_due=args.link_by_due,
            link_by_due_n=args.link_by_due_n,
            create_decks=args.create_decks,
            report_csv=args.report_csv,
            export_template_csv=args.export_template_csv,
            populate_template=args.populate_template,
            populate_output=args.populate_output,
            populate_base=args.populate_base,
            no_class_csv=args.no_class_csv,
            apply_readings=args.apply_readings,
            populate_from_db=args.populate_from_db,
            presentation_id=args.presentation_id,
            clear_semester_data=args.clear_semester_data,
        )
        return 0
    except Exception as e:
        import traceback
        print(f"Error: {e}")
        traceback.print_exc()
        return 2


if __name__ == '__main__':
    sys.exit(main())
