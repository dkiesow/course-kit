#!/usr/bin/env python3
import openpyxl

print("=" * 100)
print("DETAILED CALENDAR COMPARISON")
print("=" * 100)

# Load both files
exported = openpyxl.load_workbook('/private/tmp/SP2026_calendar.xlsx')
original = openpyxl.load_workbook('/Users/kiesowd/Library/Mobile Documents/com~apple~CloudDocs/Documents/Mizzou/Spring 2026/J4734/4734-master.xlsx')

ws_exp = exported.active
ws_orig = original.active

# Show structure
print(f"\nSTRUCTURE:")
print(f"  Exported: {ws_exp.max_row} rows x {ws_exp.max_column} columns")
print(f"  Original: {ws_orig.max_row} rows x {ws_orig.max_column} columns")
print(f"  Difference: {ws_exp.max_row - ws_orig.max_row} rows")

# Original structure (first 30 rows)
print(f"\n{'='*100}")
print("ORIGINAL CALENDAR STRUCTURE (First 30 rows):")
print(f"{'='*100}")
for i in range(1, 31):
    vals = [ws_orig[f'{c}{i}'].value for c in ['A','B','C','D','E']]
    cell_a = ws_orig[f'A{i}']
    
    markers = []
    if cell_a.value and 'UNIT' in str(cell_a.value):
        markers.append('UNIT')
    if cell_a.border.top and cell_a.border.top.style:
        markers.append(f'top-border:{cell_a.border.top.style}')
    if cell_a.fill.start_color:
        try:
            if cell_a.fill.start_color.rgb != '00000000':
                markers.append(f'fill:{cell_a.fill.start_color.rgb}')
        except:
            pass
    
    marker_str = f" [{', '.join(markers)}]" if markers else ""
    
    v0 = str(vals[0])[:18] if vals[0] else ""
    v1 = str(vals[1])[:25] if vals[1] else ""
    v2 = str(vals[2])[:18] if vals[2] else ""
    v3 = str(vals[3])[:18] if vals[3] else ""
    v4 = str(vals[4])[:18] if vals[4] else ""
    
    print(f"Row {i:2d}{marker_str:30s}: {v0:18s} | {v1:25s} | {v2:18s} | {v3:18s} | {v4:18s}")

# Exported structure (first 30 rows)
print(f"\n{'='*100}")
print("EXPORTED CALENDAR STRUCTURE (First 30 rows):")
print(f"{'='*100}")
for i in range(1, min(31, ws_exp.max_row + 1)):
    vals = [ws_exp[f'{c}{i}'].value for c in ['A','B','C','D','E']]
    cell_a = ws_exp[f'A{i}']
    
    markers = []
    if cell_a.value and 'UNIT' in str(cell_a.value):
        markers.append('UNIT')
    if cell_a.border.top and cell_a.border.top.style:
        markers.append(f'top-border:{cell_a.border.top.style}')
    if cell_a.fill.start_color:
        try:
            if hasattr(cell_a.fill.start_color, 'rgb') and cell_a.fill.start_color.rgb != '00000000':
                markers.append(f'fill:{cell_a.fill.start_color.rgb}')
        except:
            pass
    
    marker_str = f" [{', '.join(markers)}]" if markers else ""
    
    v0 = str(vals[0])[:18] if vals[0] else ""
    v1 = str(vals[1])[:25] if vals[1] else ""
    v2 = str(vals[2])[:18] if vals[2] else ""
    v3 = str(vals[3])[:18] if vals[3] else ""
    v4 = str(vals[4])[:18] if vals[4] else ""
    
    print(f"Row {i:2d}{marker_str:30s}: {v0:18s} | {v1:25s} | {v2:18s} | {v3:18s} | {v4:18s}")

print(f"\n{'='*100}")
print("KEY OBSERVATIONS:")
print(f"{'='*100}")

# Check if original has header rows
has_header = False
for i in range(1, 8):
    cell = ws_orig[f'A{i}']
    if cell.value and ('JOURN' in str(cell.value) or 'Monday' in str(cell.value)):
        has_header = True
        break

if has_header:
    print("\n✓ Original has COURSE HEADER rows (rows 1-6 with course info)")
    print("✗ Exported starts directly with UNIT ONE - missing header rows")
else:
    print("\n? Need to check original structure more carefully")

# Count UNIT rows
unit_rows_orig = [i for i in range(1, ws_orig.max_row + 1) if ws_orig[f'A{i}'].value and 'UNIT' in str(ws_orig[f'A{i}'].value)]
unit_rows_exp = [i for i in range(1, ws_exp.max_row + 1) if ws_exp[f'A{i}'].value and 'UNIT' in str(ws_exp[f'A{i}'].value)]

print(f"\nUNIT rows in original: {unit_rows_orig}")
print(f"UNIT rows in exported: {unit_rows_exp}")

# Check for lecture numbers
lecture_markers = ['ONE', 'TWO', 'THREE', 'FOUR', 'FIVE']
for marker in lecture_markers:
    found_orig = any(ws_orig[f'C{i}'].value == marker for i in range(1, ws_orig.max_row + 1))
    found_exp = any(ws_exp[f'C{i}'].value == marker for i in range(1, ws_exp.max_row + 1))
    print(f"Lecture '{marker}': Original={found_orig}, Exported={found_exp}")

print(f"\n{'='*100}")
